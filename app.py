import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import date, datetime
import json
import toml
import os
import uuid
import io
from pdf_gen import generate_pdf_report, generate_blank_pdf

# Módulos de visualización (carga lazy para no bloquear inicio)
# Módulos de visualización (carga lazy para no bloquear inicio)
try:
    from analytics import render_analytics, load_evaluaciones_df
except ImportError:
    render_analytics = None
    load_evaluaciones_df = None

try:
    from genogram import generate_genogram_dot
except ImportError:
    generate_genogram_dot = None

try:
    from ecomap import generate_ecomap_dot
except ImportError:
    generate_ecomap_dot = None

def generate_clinical_narrative(members, active_risks, nivel, programa):
    """
    Genera un análisis clínico automatizado basado en los datos de la ficha.
    """
    total_miembros = len(members)
    indices = [m for m in members if m.get("Resp", False) or str(m.get("Resp", "")).upper() == "TRUE"]
    nombre_indice = indices[0].get("Nombre y Apellidos", "N/A") if indices else "No especificado"
    
    # Análisis de Riesgos Críticos
    riesgos_t1 = [k for k in active_risks if k.startswith('t1_') and active_risks[k]]
    riesgos_t2 = [k for k in active_risks if k.startswith('t2_') and active_risks[k]]
    
    narrativa = f"### 📑 Informe de Análisis Familiar Consolidado\n\n"
    narrativa += f"**Identificación**: Familia evaluada bajo el programa **{programa}**. "
    narrativa += f"La persona índice es **{nombre_indice}**. El núcleo familiar está compuesto por **{total_miembros}** integrantes.\n\n"
    
    narrativa += f"#### 1. Clasificación del Riesgo Familiar\n"
    color_nivel = "🔴" if "ALTO" in nivel else "🟡" if "MEDIO" in nivel else "🟢"
    narrativa += f"{color_nivel} **Nivel Detectado**: {nivel}\n"
    
    if riesgos_t1:
        labels_t1 = [RISK_LABELS.get(r, r) for r in riesgos_t1]
        narrativa += f"- **Alerta Crítica**: Se detectan riesgos de alta complejidad (T1), incluyendo: {', '.join(labels_t1)}.\n"
    elif riesgos_t2:
        narrativa += f"- **Alerta Moderada**: Se observan factores protectores debilitados con presencia de riesgos T2.\n"
    else:
        narrativa += f"- **Estado**: La familia mantiene un equilibrio funcional, aunque requiere monitoreo preventivo.\n"

    narrativa += f"\n#### 2. Interpretación del Genogram y Ecomapa\n"
    narrativa += f"- **Estructura**: La dinámica generacional sugiere "
    if any("ABUEL" in str(m.get("Parentesco", "")).upper() for m in members):
        narrativa += "una estructura de familia extensa con posible apoyo multigeneracional. "
    else:
        narrativa += "una estructura nuclear con foco en la autonomía del grupo primario. "
    
    if active_risks.get('t5_redFamiliar'):
        narrativa += "Se identifica una red familiar sólida como factor protector principal. "
    else:
        narrativa += "Se observa fragilidad en los nexos internos que podría requerir intervención en comunicación. "
        
    narrativa += f"\n- **Redes Externas**: El ecomapa revela "
    if active_risks.get('t3_sinRedApoyo'):
        narrativa += "un aislamiento social significativo. Es imperativo vincular a la familia con la red comunitaria y el CESFAM. "
    else:
        narrativa += "una vinculación activa con sistemas externos, lo que facilita el proceso de intervención. "
        
    narrativa += f"\n\n#### 3. Sugerencias Prospectivas\n"
    if "ALTO" in nivel:
        narrativa += "1. Enviar caso a **Consejo Consultivo/Reunión de Equipo** para gestión de casos complejos.\n"
        narrativa += "2. Visita domiciliaria integral (VDI) en un plazo no mayor a 15 días.\n"
    elif "MEDIO" in nivel:
        narrativa += "1. Reforzar pautas de crianza y/o autocuidado según corresponda.\n"
        narrativa += "2. Seguimiento telefónico en 30 días.\n"
    else:
        narrativa += "1. Mantener controles habituales segun programa de salud.\n"
        
    return narrativa


# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(
    page_title="Ficha Riesgo Familiar",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- ESTILOS CUSTOM (SaaS PREMIUM UI) ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    /* Reset y Fondo Global */
    .stApp {
        background-color: #f8fafc; /* Gris muy claro, limpio */
        font-family: 'Inter', sans-serif;
    }
    
    /* Tipografía Global Base */
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* Ocultar elementos extra de Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Contenedor Principal "Tarjetas" */
    .main .block-container {
        max-width: 1100px;
        padding-top: 2rem;
        padding-bottom: 2rem;
        padding-left: 2rem;
        padding-right: 2rem;
    }

    /* Títulos globales */
    h1, h2, h3, h4, h5, h6 {
        color: #0f172a !important;
        font-weight: 700 !important;
        letter-spacing: -0.025em !important;
        font-family: 'Inter', sans-serif !important;
    }

    /* Botones primarios */
    button[kind="primary"] {
        background: linear-gradient(135deg, #3b82f6 0%, #4f46e5 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.25) !important;
        transition: all 0.2s ease !important;
    }
    
    button[kind="primary"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 16px rgba(59, 130, 246, 0.35) !important;
    }

    /* Botones secundarios */
    button[kind="secondary"] {
        background: white !important;
        color: #334155 !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 12px !important;
        font-weight: 500 !important;
        box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05) !important;
        transition: all 0.2s ease !important;
    }

    button[kind="secondary"]:hover {
        background: #f8fafc !important;
        border-color: #94a3b8 !important;
        color: #0f172a !important;
    }

    /* Inputs y Selects */
    .stTextInput input, .stDateInput input, .stNumberInput input, .stSelectbox div[data-baseweb="select"] {
        background-color: white !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 10px !important;
        color: #1e293b !important;
        transition: border-color 0.2s ease !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.02) !important;
    }
    
    .stTextInput input:focus, .stSelectbox div[data-baseweb="select"]:focus-within {
        border-color: #3b82f6 !important;
        box-shadow: 0 0 0 1px #3b82f6 !important;
    }

    /* Sidebar - Diseño General Limpio */
    [data-testid="stSidebar"] {
        background-color: white !important;
        border-right: 1px solid #f1f5f9;
        box-shadow: 4px 0 10px rgba(0,0,0,0.02);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: transparent;
        border-radius: 8px 8px 0 0;
        gap: 4px;
        padding-top: 10px;
        padding-bottom: 10px;
        color: #64748b;
        font-weight: 600;
        font-family: 'Inter', sans-serif;
    }
    .stTabs [aria-selected="true"] {
        color: #3b82f6 !important;
        border-bottom: 2px solid #3b82f6 !important;
    }

    /* Tarjetas de Ficha y Expander (Se aplicará en combinación con st.container) */
    .streamlit-expanderHeader {
        background-color: #f8fafc !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        color: #334155 !important;
        border: 1px solid #e2e8f0 !important;
        font-family: 'Inter', sans-serif !important;
    }
    
    div[data-testid="stExpander"] {
        border: none !important;
        box-shadow: none !important;
        background: transparent !important;
    }
    
    div[data-testid="stExpanderDetails"] {
        border: 1px solid #e2e8f0;
        border-top: none;
        border-radius: 0 0 10px 10px;
        padding: 20px;
        background: white;
    }

    /* Checkboxes */
    .stCheckbox label span {
        font-size: 0.9rem;
        color: #334155;
        font-weight: 500;
        font-family: 'Inter', sans-serif;
    }

</style>
""", unsafe_allow_html=True)

# --- CONSTANTES ---
SHEET_URL = "https://docs.google.com/spreadsheets/d/1JjYw2W6c-N2swGPuIHbz0CU7aDhh1pA-6VH1WuXV41w/edit"

PARENTESCO_OPTIONS = [
    "Jefe/a de Hogar",
    "Cónyuge/Pareja",
    "Hijo/a",
    "Padre/Madre",
    "Hermano/a",
    "Otro familiar",
    "Cuidador/a",
    "Auto-evaluación (propio)"
]

# Opciones de parentesco para la tabla de grupo familiar
PARENTESCO_FAMILIA_OPTIONS = [
    "Jefe/a de Hogar",
    "Cónyuge/Pareja",
    "Hijo/a",
    "Hijo/a (Gemelo Fraterno)",
    "Hijo/a (Gemelo Idéntico)",
    "Padre/Madre",
    "Hermano/a",
    "Abuelo/a",
    "Nieto/a",
    "Tío/a",
    "Sobrino/a",
    "Hijo/a Adoptivo/a",
    "Otro familiar",
    "No familiar"
]

TIPO_UNION_OPTIONS = ["Casados", "Convivencia", "Separados", "Divorciados"]

# Opciones de Autoidentificación Étnica (INE Chile - Censo 2017)
PUEBLO_ORIGINARIO_OPTIONS = [
    "Ninguno",
    "Mapuche",
    "Aymara",
    "Rapa Nui",
    "Atacameño (Lickanantay)",
    "Quechua",
    "Colla",
    "Diaguita",
    "Kawésqar",
    "Yagán",
    "Changos",
    "Afrodescendiente",
    "Otro (especifique en Observaciones)"
]

def generate_family_id():
    """Genera un ID único para la familia en formato FAM-AAAAMMDD-XXXX."""
    today = date.today().strftime('%Y%m%d')
    short_uuid = uuid.uuid4().hex[:6].upper()
    return f"FAM-{today}-{short_uuid}"

def generate_incremental_eval_id(familia_apellido=""):
    """
    Genera el próximo ID de Evaluación en formato incremental EVA-NNN-FAM-XXX.
    Lee la hoja 'Evaluaciones' de Google Sheets, busca el mayor número usado
    y retorna el próximo ID con las 3 letras del apellido de la familia.
    Ejemplo: EVA-001-FAM-ORT (para Familia Ortiz)
    Si no hay registros o hay error, retorna 'EVA-001-FAM-XXX'.
    """
    # Extraer prefijo del apellido (3 letras en mayúsculas, sin tilde)
    def clean_prefix(apellido):
        if not apellido or not apellido.strip():
            return "XXX"
        import unicodedata
        s = unicodedata.normalize('NFD', apellido.strip())
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')  # quita tildes
        s = ''.join(c for c in s if c.isalpha())  # solo letras
        return s[:3].upper() if len(s) >= 3 else s.upper().ljust(3, 'X')

    prefix = clean_prefix(familia_apellido)

    try:
        client = get_google_sheet_client()
        if not client:
            return f"EVA-001-FAM-{prefix}"
        spreadsheet = client.open_by_url(SHEET_URL)
        ws = spreadsheet.worksheet("Evaluaciones")
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return f"EVA-001-FAM-{prefix}"
        # Buscar columna "ID Evaluación" en la cabecera
        headers = all_vals[0]
        try:
            id_col = headers.index("ID Evaluación")
        except ValueError:
            id_col = 0  # fallback: primera columna
        # Recopilar todos los números usados en IDs con formato EVA-NNN-...
        import re
        max_num = 0
        for row in all_vals[1:]:
            if len(row) > id_col:
                cell = str(row[id_col]).strip()
                m = re.match(r'^EVA-(\d+)', cell)
                if m:
                    num = int(m.group(1))
                    if num > max_num:
                        max_num = num
        return f"EVA-{max_num + 1:03d}-FAM-{prefix}"
    except Exception:
        return f"EVA-001-FAM-{prefix}"

PROGRAMA_OPTIONS = [
    # Profesionales
    "Médico Cirujano",
    "Cirujano Dentista",
    "Enfermera / Enfermero",
    "Matrona / Matrón",
    "Kinesiólogo",
    "Nutricionista",
    "Psicólogo",
    "Fonoaudióloga",
    "Terapeuta Ocupacional",
    "Asistente Social / Trabajadora Social",
    "Químico Farmacéutico",
    "Tecnólogo Médico",
    # Técnicos de Salud
    "TENS (Técnico en Enfermería de Nivel Superior)",
    "Técnico en Odontología / Paramédico Dental",
    "Auxiliar Paramédico",
    # Ciclo Vital y Transversales
    "Salud Infantil",
    "Salud del Adolescente",
    "Salud del Adulto y Adulto Mayor",
    "Salud de la Mujer (Sexual y Reproductiva)",
    "Salud Mental",
    "Salud Odontológica (Salud Bucal)",
    # Programas Específicos y Convenios
    "Salud Cardiovascular",
    "Salud Respiratoria (IRA / ERA)",
    "Inmunizaciones (Vacunatorio)",
    "PNAC (Programa Nacional de Alimentación Complementaria)",
    "Tuberculosis",
    "IAAS (Infecciones Asociadas a la Atención de Salud)",
    "Rehabilitación Integral (RBC)",
    "Cuidados Paliativos Universales (No Oncológicos)",
    "Elige Vida Sana",
    "Espacio Amigable para Adolescentes",
    "Más Adultos Mayores Autovalentes",
    "PESPI (Programa Especial de Salud y Pueblos Indígenas)",
    "Resolutividad (Imágenes Diagnósticas, etc.)",
    "Mejoramiento de la Equidad en Salud Rural",
    # Gestión Clínica y Modelo
    "Salud Rural",
    "MAIS (Modelo de Atención Integral de Salud)",
    "ECICEP",
    "Participación Social",
    "Promoción de la Salud",
    "Calidad",
    "Capacitación",
    "Servicio de Urgencia Rural (SUR)",
]

# --- ESQUEMA DE PERMISOS RBAC ---
# Roles: 'programador', 'encargado_mais', 'jefe_sector', 'equipo_sector', 'usuario'
# Restricción REM-P7: Solo 'programador', 'encargado_mais', 'jefe_sector' (y cargos específicos definidos)

def load_users():
    """Carga la lista de usuarios desde la hoja 'usuarios' en Google Sheets."""
    try:
        client = get_google_sheet_client()
        if not client: return pd.DataFrame()
        sh = client.open_by_url(SHEET_URL)
        ws = sh.worksheet("usuarios")
        data = ws.get_all_values()
        if len(data) > 1:
            return pd.DataFrame(data[1:], columns=data[0])
    except Exception as e:
        st.error(f"Error cargando usuarios: {e}")
    return pd.DataFrame()

def check_access(row_data, user_info):
    """
    Verifica si el usuario actual tiene permiso para ver un registro específico.
    Filtra por Sector o por Programa/Unidad.
    """
    role = str(user_info.get('rol', '')).lower()
    cargo = str(user_info.get('cargo', '')).lower()
    user_unit = str(user_info.get('Programa/Unidad', '')).lower()
    
    # Programador y Encargados MAIS ven todo
    if role in ['programador', 'encargado_mais'] or 'mais' in cargo:
        return True
    
    # Filtro por Sector (Sol/Luna)
    user_unit_clean = user_unit.strip()
    user_cargo_clean = cargo.strip()
    reg_sector = str(row_data.get('Sector', '')).strip().lower()
    reg_unit = str(row_data.get('Programa/Unidad', '')).strip().lower()

    import re
    full_context = f"{user_unit_clean} {user_cargo_clean}"

    if re.search(r'\bsol\b', full_context):
        return reg_sector == 'sol'
    if re.search(r'\bluna\b', full_context) or 'postas' in full_context:
        return reg_sector == 'luna'
        
    # Filtro por Programa (ej: Cardiovascular) - Comparación flexible
    if user_unit_clean:
        if user_unit_clean in reg_unit:
            return True
        
    return False

def can_download_rem(user_info):
    """Verifica si el usuario puede descargar el reporte REM-P7."""
    role = str(user_info.get('rol', '')).lower()
    cargo = str(user_info.get('cargo', '')).lower()
    return role in ['programador', 'encargado_mais'] or 'jefe' in cargo or 'mais' in cargo or 'encargado' in cargo

RISK_LABELS = {
    't1_vif': 'Familia con VIF (física, psicológica, sexual, económica)',
    't1_drogas': 'Consumo problema de drogas o dependencia',
    't1_alcohol': 'Consumo problema de alcohol (AUDIT > 13)',
    't1_saludMentalDescomp': 'Patología salud mental descompensada o sin TTO',
    't1_abusoSexual': 'Abuso sexual (sufrido por algún integrante)',
    't1_riesgoBiopsicoGrave': 'Adulto mayor y/o niño/a en riesgo biopsicosocial grave',
    't1_epsaRiesgo': 'Pauta EPSA (ChCC) con riesgo',
    't1_vulnerabilidadExtrema': 'Vulnerabilidad socioeconómica extrema (indigencia)',
    't1_trabajoInfantil': 'Trabajo infantil en niños < 14 años',
    't2_enfermedadGrave': 'Enfermedad grave o terminal integrante',
    't2_altoRiesgoHosp': 'Paciente con alto riesgo de hospitalizar',
    't2_discapacidad': 'Discapacidad física y/o mental (Bartel 35 o menos)',
    't2_saludMentalLeve': 'Patología de salud mental leve o moderada',
    't2_judicial': 'Conflictos o problemas con la justicia',
    't2_rolesParentales': 'Incumplimiento de roles parentales',
    't2_sobrecargaCuidador': 'Sobrecarga del cuidador principal',
    't2_conflictosSeveros': 'Conflictos familiares severos o crisis de comunicación',
    't2_adultosRiesgo': 'Adultos en riesgo biopsicosocial a cargo de niños',
    't3_patologiaCronica': 'Patología crónica descompensada sintomática',
    't3_discapacidadLeve': 'Miembro con discapacidad leve/moderada (40-55pts)',
    't3_rezago': 'Rezago desarrollo psicomotor',
    't3_madreAdolescente': 'Madre adolescente',
    't3_duelo': 'Duelo reciente (pérdida de integrante significativo)',
    't3_sinRedApoyo': 'Ausencia o escasa red de apoyo social/familiar',
    't3_cesantia': 'Cesantía de más de 1 mes del proveedor',
    't3_vulneNoExtrema': 'Vulnerabilidad socioeconómica no extrema',
    't3_precariedadLaboral': 'Precariedad laboral (temporal/honorarios)',
    't3_hacinamiento': 'Hacinamiento (2.5+ personas por dormitorio)',
    't3_entornoInseguro': 'Entorno inseguro (delincuencia)',
    't3_adultoSolo': 'Adulto mayor que vive solo',
    't3_desercionEscolar': 'Deserción o fracaso escolar',
    't3_analfabetismo': 'Analfabetismo padre/madre/cuidador',
    't3_escolaridadIncompleta': 'Escolaridad básica incompleta padres',
    't3_dificultadAcceso': 'Dificultad de acceso a servicios',
    't4_monoparental': 'Hogar monoparental',
    't4_riesgoCardio': 'Riesgo cardiovascular (tabaco, obesidad)',
    't4_contaminacion': 'Foco contaminación ambiental cercano',
    't4_higiene': 'Deficiencia hábitos higiene',
    't4_sinRecreacion': 'Ausencia prácticas recreación',
    't4_sinEspaciosSeguros': 'Ausencia espacios seguros recreación',
    't4_endeudamiento': 'Endeudamiento familiar elevado (>40%)',
    't4_serviciosIncompletos': 'Servicios básicos incompletos/inadecuados',
    't5_lactancia': 'Lactancia materna exclusiva/complementaria',
    't5_habitos': 'Hábitos saludables (actividad física, alim.)',
    't5_redesSociales': 'Presencia redes sociales/comunitarias',
    't5_redFamiliar': 'Presencia red apoyo familiar',
    't5_comunicacion': 'Habilidades comunicacionales (afecto)',
    't5_recursosSuficientes': 'Recursos socioeconómicos suficientes',
    't5_resiliencia': 'Resiliencia (sobreponerse a crisis)',
    't5_viviendaAdecuada': 'Vivienda adecuada'
}


def export_rem_p7_excel(n_inscritas_sol=0, n_inscritas_luna=0):
    """
    Genera un archivo Excel con el formato oficial REM-P7.
    Retorna un objeto BytesIO listo para st.download_button.
    Si el usuario es Encargado de Postas, genera hojas por cada posta.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "Instala openpyxl: pip install openpyxl"

    # --- Leer datos de Sheets ---
    client = get_google_sheet_client()
    if not client:
        return None, "Error de conexión."
    spreadsheet = client.open_by_url(SHEET_URL)

    try:
        ws_eval = spreadsheet.worksheet("Evaluaciones")
        ev_data = ws_eval.get_all_values()
        if len(ev_data) > 1:
            df_eval = pd.DataFrame(ev_data[1:], columns=ev_data[0])
        else:
            df_eval = pd.DataFrame(columns=["ID Evaluación", "Sector", "Nivel", "Establecimiento",
                                            "egreso_alta","egreso_traslado",
                                            "egreso_derivacion","egreso_abandono"])
    except:
        df_eval = pd.DataFrame()

    try:
        ws_plan = spreadsheet.worksheet("Planes de Intervención")
        pl_data = ws_plan.get_all_values()
        df_plan = pd.DataFrame(pl_data[1:], columns=pl_data[0]) if len(pl_data) > 1 else pd.DataFrame(columns=["ID Evaluación"])
    except:
        df_plan = pd.DataFrame()

    # --- Helper para generar el contenido de una hoja ---
    def write_rem_sheet(ws, df_eval_in, df_plan_in, title_suffix="CONSOLIDADO", n_ins_sol=0, n_ins_luna=0):
        # Filtros locales para esta hoja
        def cnt(df, sector, nivel=None):
            if df.empty or "Sector" not in df.columns: return 0
            m = df["Sector"].str.strip().str.lower() == sector.lower()
            if nivel: m &= df["Nivel"].str.strip().str.upper() == nivel.upper()
            return int(m.sum())

        def cnt_bool(df, sector, col):
            if df.empty or col not in df.columns: return 0
            m = (df["Sector"].str.strip().str.lower() == sector.lower()) & \
                (df[col].astype(str).str.strip().str.upper().isin(["TRUE","1","VERDADERO"]))
            return int(m.sum())

        ids_sol = set(); ids_luna = set()
        if not df_plan_in.empty and "ID Evaluación" in df_plan_in.columns:
            ids_sol  = {i for i in df_plan_in["ID Evaluación"].unique()
                        if not df_eval_in[df_eval_in["ID Evaluación"]==i].empty
                        and df_eval_in[df_eval_in["ID Evaluación"]==i].iloc[0].get("Sector","").strip().lower()=="sol"}
            ids_luna = {i for i in df_plan_in["ID Evaluación"].unique()
                        if not df_eval_in[df_eval_in["ID Evaluación"]==i].empty
                        and df_eval_in[df_eval_in["ID Evaluación"]==i].iloc[0].get("Sector","").strip().lower()=="luna"}

        sol_ev = cnt(df_eval_in,"Sol"); luna_ev = cnt(df_eval_in,"Luna")
        sol_b  = cnt(df_eval_in,"Sol","RIESGO BAJO");   luna_b  = cnt(df_eval_in,"Luna","RIESGO BAJO")
        sol_m  = cnt(df_eval_in,"Sol","RIESGO MEDIO");  luna_m  = cnt(df_eval_in,"Luna","RIESGO MEDIO")
        sol_a  = cnt(df_eval_in,"Sol","RIESGO ALTO");   luna_a  = cnt(df_eval_in,"Luna","RIESGO ALTO")
        sol_cp = len(ids_sol); luna_cp = len(ids_luna)

        def sin_plan_local(nivel_str, ids_con, sector):
            ids_nivel = set(df_eval_in[df_eval_in["Nivel"].str.strip().str.upper()==nivel_str.upper()]["ID Evaluación"]) \
                        if not df_eval_in.empty else set()
            ids_sector = set(df_eval_in[df_eval_in["Sector"].str.strip().str.lower()==sector.lower()]["ID Evaluación"]) \
                         if not df_eval_in.empty else set()
            return max(len(ids_nivel & ids_sector) - len(ids_con & ids_nivel & ids_sector), 0)

        sol_sb   = sin_plan_local("RIESGO BAJO",ids_sol,"sol")
        sol_sm   = sin_plan_local("RIESGO MEDIO",ids_sol,"sol")
        sol_sa   = sin_plan_local("RIESGO ALTO",ids_sol,"sol")
        luna_sb  = sin_plan_local("RIESGO BAJO",ids_luna,"luna")
        luna_sm  = sin_plan_local("RIESGO MEDIO",ids_luna,"luna")
        luna_sa  = sin_plan_local("RIESGO ALTO",ids_luna,"luna")

        eg_types = ["egreso_alta","egreso_traslado","egreso_derivacion","egreso_abandono"]
        sol_egs   = {c: cnt_bool(df_eval_in, "sol", c) for c in eg_types}
        luna_egs  = {c: cnt_bool(df_eval_in, "luna", c) for c in eg_types}
        sol_eg_t  = sum(sol_egs.values()); luna_eg_t = sum(luna_egs.values())

        # Estilos
        DARK_BLUE   = PatternFill("solid", fgColor="1F3864"); YELLOW = PatternFill("solid", fgColor="FFD966")
        CELESTE     = PatternFill("solid", fgColor="BDD7EE"); CELESTE_LT = PatternFill("solid", fgColor="DEEAF1")
        CELESTE_MID = PatternFill("solid", fgColor="9DC3E6"); BOLD_WHITE = Font(bold=True, color="FFFFFF", size=11)
        BOLD_DARK   = Font(bold=True, color="1F3864", size=10); NORMAL = Font(size=9, color="000000")
        THIN = Side(style="thin", color="B8CCE4"); THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True); LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
        WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

        def set_cell(row, col, value, fill=None, font=None, align=None, border=True):
            cell = ws.cell(row=row, column=col, value=value)
            if fill: cell.fill = fill
            if font: cell.font = font
            if align: cell.alignment = align
            if border: cell.border = THIN_BORDER
            return cell

        def merge_row(row, c1, c2, value, fill=None, font=None, align=None):
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = ws.cell(row=row, column=c1, value=value)
            if fill: cell.fill = fill
            if font: cell.font = font
            if align: cell.alignment = align
            cell.border = THIN_BORDER

        r = 1
        merge_row(r, 1, 11, f"REM-P7. FAMILIAS EN CONTROL SALUD FAMILIAR - {title_suffix}", DARK_BLUE, BOLD_WHITE, CENTER)
        r += 1
        merge_row(r, 1, 11, f"CESFAM Cholchol | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", CELESTE, BOLD_DARK, CENTER)
        
        # SECCIÓN A
        r += 1
        merge_row(r, 1, 11, "SECCIÓN A. CLASIFICACIÓN FAMILIas URBANO (Sector Sol)", YELLOW, BOLD_DARK, LEFT)
        r += 1
        cols_A = ["Clasificación", "TOTAL", "Sector Sol", "", "", "", "", "", "", ""]
        for ci, h in enumerate(cols_A, 1): set_cell(r, ci, h, CELESTE, BOLD_DARK, CENTER)
        
        secA_data = [
            ("N° Familias inscritas", n_ins_sol, n_ins_sol),
            ("N° Familias evaluadas", sol_ev, sol_ev),
            ("N° Riesgo bajo", sol_b, sol_b),
            ("N° Riesgo medio", sol_m, sol_m),
            ("N° Riesgo alto", sol_a, sol_a),
        ]
        for label, total, s1 in secA_data:
            r += 1
            set_cell(r, 1, label, CELESTE_LT, NORMAL, LEFT)
            set_cell(r, 2, total, CELESTE, BOLD_DARK, CENTER)
            set_cell(r, 3, s1, WHITE_FILL, NORMAL, CENTER)

        # SECCIÓN A.1
        r += 1
        merge_row(r, 1, 11, "SECCIÓN A.1 CLASIFICACIÓN FAMILIas RURAL (Sector Luna)", YELLOW, BOLD_DARK, LEFT)
        r += 1
        for ci, h in enumerate(cols_A, 1): set_cell(r, ci, h.replace("Sol", "Luna"), CELESTE, BOLD_DARK, CENTER)
        secA1_data = [
            ("N° Familias inscritas", n_ins_luna, n_ins_luna),
            ("N° Familias evaluadas", luna_ev, luna_ev),
            ("N° Riesgo bajo", luna_b, luna_b),
            ("N° Riesgo medio", luna_m, luna_m),
            ("N° Riesgo alto", luna_a, luna_a),
        ]
        for label, total, s1 in secA1_data:
            r += 1
            set_cell(r, 1, label, CELESTE_LT, NORMAL, LEFT)
            set_cell(r, 2, total, CELESTE, BOLD_DARK, CENTER)
            set_cell(r, 3, s1, WHITE_FILL, NORMAL, CENTER)

        # SECCIÓN B
        r += 1
        merge_row(r, 1, 11, "SECCIÓN B. INTERVENCIÓN URBANO Y RURAL", YELLOW, BOLD_DARK, LEFT)
        r += 1
        cols_B = ["Intervención", "", "TOTAL", "Sol", "Luna", "", "", "", "", "", ""]
        for ci, h in enumerate(cols_B, 1): set_cell(r, ci, h, CELESTE, BOLD_DARK, CENTER)
        
        secB_data = [
            ("N° Fam con plan", "", sol_cp+luna_cp, sol_cp, luna_cp),
            ("N° Fam sin plan", "Riesgo bajo", sol_sb+luna_sb, sol_sb, luna_sb),
            ("", "Riesgo medio", sol_sm+luna_sm, sol_sm, luna_sm),
            ("", "Riesgo alto", sol_sa+luna_sa, sol_sa, luna_sa),
            ("N° Fam egresadas", "Total egresos", sol_eg_t+luna_eg_t, sol_eg_t, luna_eg_t),
        ]
        for l1, l2, tot, s1, s2 in secB_data:
            r += 1
            set_cell(r, 1, l1, CELESTE_LT, NORMAL, LEFT)
            set_cell(r, 2, l2, CELESTE_MID, NORMAL, LEFT)
            set_cell(r, 3, tot, CELESTE, BOLD_DARK, CENTER)
            set_cell(r, 4, s1, WHITE_FILL, NORMAL, CENTER)
            set_cell(r, 5, s2, WHITE_FILL, NORMAL, CENTER)

        col_widths = [35, 18, 10, 10, 10, 8, 8, 8, 8, 8, 8]
        for i, w in enumerate(col_widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    # --- Lógica Principal ---
    user_info = st.session_state.get('user_info', {})
    cargo = str(user_info.get('cargo', '')).lower()
    is_posta_role = 'encargado' in cargo and 'postas' in cargo

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "CONSOLIDADO"
    
    write_rem_sheet(ws_main, df_eval, df_plan, "CONSOLIDADO", n_inscritas_sol, n_inscritas_luna)

    if is_posta_role:
        # Generar hojas por establecimiento para Sector Luna
        if not df_eval.empty and 'Establecimiento' in df_eval.columns:
            est_list = [e for e in df_eval['Establecimiento'].unique() if e and str(e).strip()]
            for est_name in est_list:
                clean_name = str(est_name).strip()
                if clean_name.lower() == "cesfam cholchol": continue
                
                # Crear hoja limpia para el establecimiento
                ws_est = wb.create_sheet(title=clean_name[:31])
                df_ev_est = df_eval[df_eval['Establecimiento'] == est_name]
                write_rem_sheet(ws_est, df_ev_est, df_plan, clean_name.upper(), 0, len(df_ev_est))

    # Guardar en buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# --- GOOGLE SHEETS CONNECTION ---
def get_google_sheet_client():
    try:
        secrets = st.secrets["gcp_service_account"]
        creds_dict = {
            "type": secrets["type"],
            "project_id": secrets["project_id"],
            "private_key_id": secrets["private_key_id"],
            "private_key": secrets["private_key"],
            "client_email": secrets["client_email"],
            "client_id": secrets["client_id"],
            "auth_uri": secrets["auth_uri"],
            "token_uri": secrets["token_uri"],
            "auth_provider_x509_cert_url": secrets["auth_provider_x509_cert_url"],
            "client_x509_cert_url": secrets["client_x509_cert_url"],
            "universe_domain": secrets["universe_domain"]
        }
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        st.error(f"Error conectando a Google Sheets: {e}")
        return None

@st.cache_data(ttl=300)
def get_all_ruts_mapping():
    """Retorna un dict {rut: (familia, id_eval)} de todos los integrantes de la BD para validación."""
    try:
        client = get_google_sheet_client()
        if not client: return {}
        sh = client.open_by_url(SHEET_URL)
        ws = sh.worksheet("Evaluaciones")
        data = ws.get_all_values()
        if len(data) <= 1: return {}
        
        df = pd.DataFrame(data[1:], columns=data[0])
        rut_map = {}
        for _, row in df.iterrows():
            fid = row.get("ID Evaluación", "Desconocido")
            fam = row.get("Familia", "Desconocida")
            gf_json = row.get("Grupo Familiar JSON", "[]")
            try:
                gf = json.loads(gf_json)
                for member in gf:
                    rut = str(member.get("RUT", "")).strip().upper()
                    if rut and rut != "S/R" and rut != "":
                        rut_map[rut] = (fam, fid)
            except:
                continue
        return rut_map
    except Exception as e:
        st.error(f"Error cargando mapeo de RUTs: {e}")
        return {}


def get_or_create_worksheet(spreadsheet, title, headers=None):
    """Obtiene una hoja por nombre, la crea si no existe y opcionalmente pone encabezados."""
    try:
        ws = spreadsheet.worksheet(title)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=title, rows="1000", cols="50")
        if headers:
            ws.append_row(headers)
    return ws

def log_audit_event(user_info, action, details="", eval_id=None):
    """Registra un evento de auditoría en la hoja 'Auditoría' de Google Sheets."""
    try:
        client = get_google_sheet_client()
        if not client:
            return
        spreadsheet = client.open_by_url(SHEET_URL)
        headers = ["Timestamp", "Usuario", "Cargo", "Acción", "Detalles", "ID Evaluación"]
        worksheet = get_or_create_worksheet(spreadsheet, "Auditoría", headers)
        
        # Obtener fecha y hora actual
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        row = [
            timestamp,
            user_info.get('usuario', 'Desconocido'),
            user_info.get('cargo', 'Desconocido'),
            action,
            details,
            str(eval_id) if eval_id else ""
        ]
        worksheet.append_row(row)
    except Exception as e:
        # Fallo silencioso en auditoría para no bloquear la experiencia de usuario
        print(f"Error registrando auditoría: {e}")


def search_record(id_eval):
    client = get_google_sheet_client()
    if not client:
        return None
    try:
        spreadsheet = client.open_by_url(SHEET_URL)
        worksheet = get_or_create_worksheet(spreadsheet, "Evaluaciones")

        # Get all values (list of lists) to avoid duplicate header errors
        all_values = worksheet.get_all_values()
        if not all_values:
            return None
            
        headers = all_values[0]
        try:
            id_idx = headers.index("ID Evaluación")
        except ValueError:
            return None
            
        for row in all_values[1:]:
            if len(row) > id_idx:
                val = row[id_idx]
                if str(val) == str(id_eval):
                    record = {}
                    for i, h in enumerate(headers):
                        if h and i < len(row):
                             record[h] = row[i]
                    
                    # VALIDACIÓN RBAC: Verificar si el usuario actual tiene permiso para ver este registro
                    if 'authenticated' in st.session_state and st.session_state.authenticated:
                        if not check_access(record, st.session_state.user_info):
                            st.error("🚫 No tiene permisos para acceder a este registro (Restricción de Sector/Unidad).")
                            return None
                    
                    return record
        return None
    except Exception as e:
        st.error(f"Error buscando en Google Sheets: {e}")
        return None


def load_record_into_state(record):
    mapping = {
        'Rep Sector': 'comp_rep_sector',
        'Familia Comp': 'comp_familia',
        'Dir Comp': 'comp_dir',
        'Rep Familia': 'comp_rep_fam',
        'RUT Rep': 'comp_rut',
        'Firma Funcionario': 'sig_func',
        'Firma Beneficiario': 'sig_benef',
        'Firma Equipo': 'sig_equipo',
        'Firma Jefe': 'sig_jefe',
        'Firma Evaluador': 'sig_evaluador_input',
        'Evaluador': 'evaluadorName',
        'Parentesco': 'parentesco',
        'Programa/Unidad': 'programa_unidad',
        'Sector': 'sector',
        'Tipo Unión': 'tipo_union',
        'Observaciones': 'observaciones',
    }
    
    for header, state_key in mapping.items():
        st.session_state[state_key] = record.get(header, '')

    # APGAR
    try:
        st.session_state['apgar_total'] = int(record.get('APGAR Total', 0))
        st.session_state['apgar_a1'] = int(record.get('A1', 0))
        st.session_state['apgar_a2'] = int(record.get('A2', 0))
        st.session_state['apgar_a3'] = int(record.get('A3', 0))
        st.session_state['apgar_a4'] = int(record.get('A4', 0))
        st.session_state['apgar_a5'] = int(record.get('A5', 0))
    except (ValueError, TypeError):
        st.session_state['apgar_total'] = 0
        for i in range(1, 6): st.session_state[f'apgar_a{i}'] = 0

    # Dates
    try:
        if record.get('Fecha', ''):
             st.session_state['fechaEvaluacion'] = datetime.strptime(str(record.get('Fecha', '')), '%Y-%m-%d').date()
        else:
             st.session_state['fechaEvaluacion'] = date.today()
    except:
        st.session_state['fechaEvaluacion'] = date.today()

    try:
        if record.get('Fecha Comp', ''):
             st.session_state['comp_fecha'] = datetime.strptime(str(record.get('Fecha Comp', '')), '%Y-%m-%d').date()
    except:
        pass

    try:
        if record.get('Fecha Egreso', ''):
             st.session_state['fechaEgreso'] = datetime.strptime(str(record.get('Fecha Egreso', '')), '%Y-%m-%d').date()
    except:
        st.session_state['fechaEgreso'] = None

    # Booleans (Risk Factors + Egreso)
    for key in risk_keys:
        val = record.get(key, False)
        st.session_state[key] = True if str(val).upper() in ['TRUE', '1', 'YES'] else False
        
    egreso_keys = ['egreso_alta', 'egreso_traslado', 'egreso_derivacion', 'egreso_abandono']
    for k in egreso_keys:
        val = record.get(k, False)
        st.session_state[k] = True if str(val).upper() in ['TRUE', '1', 'YES'] else False

    # Dynamic Tables (JSON)
    try:
        fam_json = record.get('Grupo Familiar JSON', '[]')
        plan_json = record.get('Plan Intervención JSON', '[]')
        
        df_fam = pd.DataFrame(json.loads(fam_json) if fam_json else [])
        if not df_fam.empty:
            if 'Sexo' in df_fam.columns and 'Identidad de género' not in df_fam.columns:
                migration_map = {"M": "Masculino", "F": "Femenino", "G": "Gestación/Aborto"}
                df_fam['Identidad de género'] = df_fam['Sexo'].map(lambda x: migration_map.get(str(x).upper(), str(x)))
            if 'F. Nac' in df_fam.columns:
                df_fam['F. Nac'] = pd.to_datetime(df_fam['F. Nac'], errors='coerce').dt.date
        st.session_state.family_members = df_fam
        
        df_plan = pd.DataFrame(json.loads(plan_json) if plan_json else [])
        cols_date = ['Fecha Prog', 'Fecha Real', 'F. Seguimiento']
        for c in cols_date:
            if c in df_plan.columns:
                df_plan[c] = pd.to_datetime(df_plan[c], errors='coerce').dt.date
        # Backfill new tracking columns if missing (older records)
        for tracking_col in ['Estado', 'F. Seguimiento', 'Obs. Seguimiento']:
            if tracking_col not in df_plan.columns:
                df_plan[tracking_col] = None if 'Fecha' in tracking_col else ''
        
        st.session_state.intervention_plan = df_plan

        # Load seguimiento plan
        seg_json_load = record.get('Seguimiento Plan JSON', '[]')
        try:
            df_seg = pd.DataFrame(json.loads(seg_json_load) if seg_json_load else [])
            for _sc in ['Objetivo', 'Actividad', 'Estado', 'F. Seguimiento', 'Obs. Seguimiento']:
                if _sc not in df_seg.columns:
                    df_seg[_sc] = ''
            if 'F. Seguimiento' in df_seg.columns:
                df_seg['F. Seguimiento'] = pd.to_datetime(df_seg['F. Seguimiento'], errors='coerce').dt.date
            st.session_state.seguimiento_plan = df_seg
        except:
            st.session_state.seguimiento_plan = pd.DataFrame(columns=['Objetivo', 'Actividad', 'Estado', 'F. Seguimiento', 'Obs. Seguimiento'])
    except Exception as e:
        st.warning(f"Error cargando tablas: {e}")

    try:
        rel_json = record.get('Relaciones JSON', '[]')
        st.session_state.interpersonal_relations = json.loads(rel_json) if rel_json else []
    except:
        st.session_state.interpersonal_relations = []

    try:
        team_json = record.get('Equipo Salud JSON', '[]')
        try:
            df_team = pd.DataFrame(json.loads(team_json))
        except:
            df_team = pd.DataFrame(columns=["Nombre y Profesión", "Firma"])
            
        if "Nombre" in df_team.columns and "Cargo" in df_team.columns:
            df_team["Nombre y Profesión"] = df_team.apply(lambda x: f"{x['Nombre']} - {x['Cargo']}", axis=1)
            if "Firma" not in df_team.columns: df_team["Firma"] = False
            df_team = df_team[["Nombre y Profesión", "Firma"]]
            
        st.session_state.team_members = df_team if not df_team.empty else pd.DataFrame(columns=["Nombre y Profesión", "Firma"])
    except:
        pass

    return record


def save_evaluacion_to_sheet(data, headers):
    """Guarda o actualiza la evaluación en la Hoja 1 'Evaluaciones'."""
    client = get_google_sheet_client()
    if not client:
        return False, "Error de conexión."
    try:
        spreadsheet = client.open_by_url(SHEET_URL)
        worksheet = get_or_create_worksheet(spreadsheet, "Evaluaciones", headers)

        all_values = worksheet.get_all_values()
        
        # Ensure header row exists
        if not all_values:
            worksheet.append_row(headers)
            all_values = [headers]
        elif all_values[0] != headers:
            # Si los encabezados cambiaron, actualizar la primera fila
            worksheet.update(range_name="A1", values=[headers])
            all_values[0] = headers

        new_id = str(data[0]).strip()
        
        if not new_id:
             worksheet.append_row(data)
             if 'raw_analytics_df' in st.session_state: del st.session_state['raw_analytics_df']
             return True, "Registro agregado (sin ID)."

        id_col_idx = 0
        if all_values and "ID Evaluación" in all_values[0]:
             try:
                 id_col_idx = all_values[0].index("ID Evaluación")
             except:
                 id_col_idx = 0
        
        row_to_update = -1
        start_row_idx = 1 if (all_values and len(all_values) > 0) else 0
        
        for i in range(start_row_idx, len(all_values)):
            row = all_values[i]
            if len(row) > id_col_idx:
                existing_id = str(row[id_col_idx]).strip()
                if existing_id == new_id:
                    row_to_update = i + 1
                    break
        
        if row_to_update != -1:
            worksheet.update(range_name=f"A{row_to_update}", values=[data])
            if 'raw_analytics_df' in st.session_state: del st.session_state['raw_analytics_df']
            return True, f"Registro actualizado (Fila {row_to_update})."
        else:
            worksheet.append_row(data)
            if 'raw_analytics_df' in st.session_state: del st.session_state['raw_analytics_df']
            return True, "Nuevo registro agregado."
            
    except Exception as e:
        st.error(f"Error guardando en Hoja Evaluaciones: {e}")
        return False, str(e)


def migrate_eval_ids_to_new_format():
    """
    Migra todos los IDs de Evaluaci\u00f3n existentes en Google Sheets al nuevo formato EVA-NNN-FAM-XXX.
    Itera sobre la hoja 'Evaluaciones', extrae el apellido de la columna 'Familia',
    genera nuevos IDs secuenciales y actualiza las filas.
    Tambi\u00e9n actualiza referencias en Planes de Intervenci\u00f3n y Ecomapas.
    Retorna (bool, mensaje, n\u00famero de registros actualizados).
    """
    import unicodedata, re

    def clean_prefix(apellido):
        if not apellido or not str(apellido).strip():
            return "XXX"
        s = unicodedata.normalize('NFD', str(apellido).strip())
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        s = ''.join(c for c in s if c.isalpha())
        return s[:3].upper() if len(s) >= 3 else s.upper().ljust(3, 'X')

    client = get_google_sheet_client()
    if not client:
        return False, "Error de conexi\u00f3n con Google Sheets.", 0

    try:
        spreadsheet = client.open_by_url(SHEET_URL)

        # Leer hoja de Evaluaciones
        ws_eval = spreadsheet.worksheet("Evaluaciones")
        all_vals = ws_eval.get_all_values()
        if len(all_vals) < 2:
            return False, "No hay registros en la hoja Evaluaciones.", 0

        headers = all_vals[0]
        data_rows = all_vals[1:]

        # Detectar columnas clave
        id_col  = headers.index("ID Evaluaci\u00f3n") if "ID Evaluaci\u00f3n" in headers else 0
        fam_col = headers.index("Familia") if "Familia" in headers else 2

        # Leer hoja Planes de Intervenci\u00f3n
        try:
            ws_plan = spreadsheet.worksheet("Planes de Intervenci\u00f3n")
            plan_all = ws_plan.get_all_values()
            plan_headers = plan_all[0] if plan_all else []
            plan_id_col = plan_headers.index("ID Evaluaci\u00f3n") if "ID Evaluaci\u00f3n" in plan_headers else 0
        except Exception:
            ws_plan, plan_all, plan_id_col = None, [], 0

        # Leer hoja Ecomapas
        try:
            ws_eco = spreadsheet.worksheet("Ecomapas")
            eco_all = ws_eco.get_all_values()
            eco_headers = eco_all[0] if eco_all else []
            eco_id_col = eco_headers.index("ID Evaluaci\u00f3n") if "ID Evaluaci\u00f3n" in eco_headers else 0
        except Exception:
            ws_eco, eco_all, eco_id_col = None, [], 0

        # Construir mapa: ID_viejo -> ID_nuevo
        id_map = {}
        updates_eval = []
        counter = 1
        for row in data_rows:
            old_id   = str(row[id_col]).strip() if len(row) > id_col else ""
            apellido = str(row[fam_col]).strip().split()[0] if len(row) > fam_col and str(row[fam_col]).strip() else ""
            prefix   = clean_prefix(apellido)
            new_id   = f"EVA-{counter:03d}-FAM-{prefix}"
            id_map[old_id] = new_id
            updated_row = list(row)
            updated_row[id_col] = new_id
            updates_eval.append((counter + 1, updated_row))  # +1 por encabezado
            counter += 1

        # Actualizar Evaluaciones fila por fila
        for sheet_row_num, updated_row in updates_eval:
            ws_eval.update(range_name=f"A{sheet_row_num}", values=[updated_row])

        # Actualizar Planes de Intervenci\u00f3n
        if ws_plan and len(plan_all) > 1:
            for i, row in enumerate(plan_all[1:], 2):
                if len(row) > plan_id_col:
                    old_id = str(row[plan_id_col]).strip()
                    if old_id in id_map:
                        updated_row = list(row)
                        updated_row[plan_id_col] = id_map[old_id]
                        ws_plan.update(range_name=f"A{i}", values=[updated_row])

        # Actualizar Ecomapas
        if ws_eco and len(eco_all) > 1:
            for i, row in enumerate(eco_all[1:], 2):
                if len(row) > eco_id_col:
                    old_id = str(row[eco_id_col]).strip()
                    if old_id in id_map:
                        updated_row = list(row)
                        updated_row[eco_id_col] = id_map[old_id]
                        ws_eco.update(range_name=f"A{i}", values=[updated_row])

        return True, f"Migraci\u00f3n completada: {len(updates_eval)} registros actualizados.", len(updates_eval)

    except Exception as e:
        return False, f"Error durante la migraci\u00f3n: {e}", 0


def save_intervention_rows(id_eval, familia, fecha_eval, nivel, programa, parentesco, df_plan):
    """Guarda las filas del plan de intervención en la Hoja 2 'Planes de Intervención'.
    
    Primero elimina registros existentes con el mismo ID Evaluación, luego inserta las nuevas filas.
    """
    client = get_google_sheet_client()
    if not client:
        return False, "Error de conexión."
    
    plan_headers = [
        "ID Evaluación", "Familia", "Fecha Evaluación", "Nivel Riesgo",
        "Programa/Unidad", "Parentesco",
        "Objetivo", "Actividad", "Fecha Prog", "Responsable", "Fecha Real", "Evaluación",
        "Estado", "F. Seguimiento", "Obs. Seguimiento"
    ]
    
    try:
        spreadsheet = client.open_by_url(SHEET_URL)
        worksheet = get_or_create_worksheet(spreadsheet, "Planes de Intervención", plan_headers)
        
        all_values = worksheet.get_all_values()
        
        # Ensure headers
        if not all_values:
            worksheet.append_row(plan_headers)
            all_values = [plan_headers]

        # Find and delete existing rows with same ID Evaluación
        if len(all_values) > 1:
            headers_row = all_values[0]
            try:
                id_idx = headers_row.index("ID Evaluación")
            except ValueError:
                id_idx = 0

            rows_to_delete = []
            for i in range(1, len(all_values)):
                row = all_values[i]
                if len(row) > id_idx and str(row[id_idx]).strip() == str(id_eval).strip():
                    rows_to_delete.append(i + 1)  # 1-indexed sheet row

            # Delete from bottom to top to preserve row numbers
            for row_num in sorted(rows_to_delete, reverse=True):
                worksheet.delete_rows(row_num)

        # Insert new rows for each plan activity
        if df_plan is not None and not df_plan.empty:
            new_rows = []
            for _, plan_row in df_plan.iterrows():
                fecha_prog = ""
                fecha_real = ""
                try:
                    fp = plan_row.get("Fecha Prog")
                    if fp and pd.notnull(fp):
                        fecha_prog = fp.strftime('%Y-%m-%d') if hasattr(fp, 'strftime') else str(fp)
                except:
                    pass
                try:
                    fr = plan_row.get("Fecha Real")
                    if fr and pd.notnull(fr):
                        fecha_real = fr.strftime('%Y-%m-%d') if hasattr(fr, 'strftime') else str(fr)
                except:
                    pass
                    
                fecha_seg = ""
                try:
                    fs = plan_row.get("F. Seguimiento")
                    if fs and pd.notnull(fs):
                        fecha_seg = fs.strftime('%Y-%m-%d') if hasattr(fs, 'strftime') else str(fs)
                except:
                    pass

                new_rows.append([
                    str(id_eval),
                    str(familia),
                    str(fecha_eval),
                    str(nivel),
                    str(programa),
                    str(parentesco),
                    str(plan_row.get("Objetivo", "")),
                    str(plan_row.get("Actividad", "")),
                    fecha_prog,
                    str(plan_row.get("Responsable", "")),
                    fecha_real,
                    str(plan_row.get("Evaluación", "")),
                    str(plan_row.get("Estado", "")),
                    fecha_seg,
                    str(plan_row.get("Obs. Seguimiento", ""))
                ])
            
            if new_rows:
                worksheet.append_rows(new_rows, value_input_option='USER_ENTERED')
                return True, f"{len(new_rows)} actividades guardadas en Hoja 'Planes de Intervención'."
            else:
                return True, "Plan de intervención vacío, no se agregaron filas."
        else:
            return True, "Plan de intervención vacío, no se agregaron filas."

    except Exception as e:
        st.error(f"Error guardando en Hoja Planes de Intervención: {e}")
        return False, str(e)



# --- PERSISTENCIA ECOMAPA ---
def save_ecomap_to_sheet(eval_id, familia, elements):
    """
    Guarda la configuración del ecomapa en la hoja 'Ecomapas'.
    Incluye los flujos de energía del Protocolo SJ.
    """
    try:
        client = get_google_sheet_client()
        if not client: return False, "No se pudo conectar con Google."
        spreadsheet = client.open_by_url(SHEET_URL)
        
        headers = ["ID Evaluación", "Familia", "Sistemas JSON", "Flujos JSON", "Riesgos JSON", "Fecha Actualización"]
        ws = get_or_create_worksheet(spreadsheet, "Ecomapas", headers)
        
        # Serializar la configuración
        sistemas_json = json.dumps(elements.get("selected_systems", []), ensure_ascii=False)
        flujos_json = json.dumps(elements.get("system_flows", {}), ensure_ascii=False)
        riesgos_json = json.dumps(elements.get("active_risks", {}), ensure_ascii=False)
        fecha_act = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Buscar si ya existe para este ID
        all_vals = ws.get_all_values()
        row_idx = -1
        if len(all_vals) > 1:
            for i, row in enumerate(all_vals[1:], 2):
                if len(row) > 0 and str(row[0]).strip() == str(eval_id).strip():
                    row_idx = i
                    break
        
        new_row = [eval_id, familia, sistemas_json, flujos_json, riesgos_json, fecha_act]
        if row_idx != -1:
            ws.update(f"A{row_idx}:F{row_idx}", [new_row])
            return True, "Ecomapa actualizado."
        else:
            ws.append_row(new_row)
            return True, "Ecomapa guardado."
            
    except Exception as e:
        return False, f"Error ecomapa: {e}"


def update_rem_p7(n_inscritas_sol=0, n_inscritas_luna=0):
    """
    Regenera la hoja 'REM-P7' con el resumen estadístico REM-P7 Familias en Control.
    Lee todos los registros de 'Evaluaciones' y 'Planes de Intervención' para calcular los totales.
    
    Parámetros:
        n_inscritas_sol:  N° total de familias inscritas en sector Sol
        n_inscritas_luna: N° total de familias inscritas en sector Luna
    """
    client = get_google_sheet_client()
    if not client:
        return False, "Error de conexión."
    try:
        spreadsheet = client.open_by_url(SHEET_URL)

        # ------- Leer datos de evaluaciones -------
        try:
            ws_eval = spreadsheet.worksheet("Evaluaciones")
            eval_data = ws_eval.get_all_values()
        except:
            eval_data = []

        # Construir DataFrame de evaluaciones
        if len(eval_data) > 1:
            df_eval = pd.DataFrame(eval_data[1:], columns=eval_data[0])
        else:
            df_eval = pd.DataFrame(columns=["ID Evaluación", "Sector", "Nivel",
                                            "egreso_alta", "egreso_traslado",
                                            "egreso_derivacion", "egreso_abandono"])

        # ------- Leer datos de planes -------
        try:
            ws_plan = spreadsheet.worksheet("Planes de Intervención")
            plan_data = ws_plan.get_all_values()
        except:
            plan_data = []

        if len(plan_data) > 1:
            df_plan = pd.DataFrame(plan_data[1:], columns=plan_data[0])
        else:
            df_plan = pd.DataFrame(columns=["ID Evaluación"])

        # ------- Cálculos por sector -------
        def count_sector(df, sector_val, nivel=None, col="Nivel"):
            if df.empty or col not in df.columns:
                return 0
            mask = df["Sector"].str.strip().str.lower() == sector_val.lower()
            if nivel:
                mask &= df[col].str.strip().str.upper() == nivel.upper()
            return int(mask.sum())

        def count_bool_col(df, sector_val, col):
            """Cuenta filas donde col es TRUE para un sector."""
            if df.empty or col not in df.columns:
                return 0
            mask = (df["Sector"].str.strip().str.lower() == sector_val.lower()) & \
                   (df[col].str.strip().str.upper().isin(["TRUE", "1", "YES", "VERDADERO"]))
            return int(mask.sum())

        # IDs con plan de intervención (tienen filas en Hoja 2)
        ids_con_plan_sol = set()
        ids_con_plan_luna = set()
        if not df_plan.empty and "ID Evaluación" in df_plan.columns:
            for eval_id in df_plan["ID Evaluación"].unique():
                rows_eval = df_eval[df_eval["ID Evaluación"] == eval_id] if not df_eval.empty else pd.DataFrame()
                if not rows_eval.empty:
                    sector = rows_eval.iloc[0].get("Sector", "").strip().lower()
                    if sector == "sol":
                        ids_con_plan_sol.add(eval_id)
                    elif sector == "luna":
                        ids_con_plan_luna.add(eval_id)

        # Sector Sol (Sector 1)
        sol_eval          = count_sector(df_eval, "Sol")
        sol_bajo          = count_sector(df_eval, "Sol", "RIESGO BAJO")
        sol_medio         = count_sector(df_eval, "Sol", "RIESGO MEDIO")
        sol_alto          = count_sector(df_eval, "Sol", "RIESGO ALTO")
        sol_con_plan      = len(ids_con_plan_sol)
        sol_sin_bajo      = sol_bajo  - sum(1 for id_ in ids_con_plan_sol if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO BAJO")
        sol_sin_medio     = sol_medio - sum(1 for id_ in ids_con_plan_sol if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO MEDIO")
        sol_sin_alto      = sol_alto  - sum(1 for id_ in ids_con_plan_sol if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO ALTO")
        sol_egreso        = count_bool_col(df_eval, "Sol", "egreso_alta") + count_bool_col(df_eval, "Sol", "egreso_traslado") + count_bool_col(df_eval, "Sol", "egreso_derivacion") + count_bool_col(df_eval, "Sol", "egreso_abandono")
        sol_egreso_alta   = count_bool_col(df_eval, "Sol", "egreso_alta")
        sol_egreso_tras   = count_bool_col(df_eval, "Sol", "egreso_traslado")
        sol_egreso_deriv  = count_bool_col(df_eval, "Sol", "egreso_derivacion")
        sol_egreso_aban   = count_bool_col(df_eval, "Sol", "egreso_abandono")

        # Sector Luna (Sector 2)
        luna_eval         = count_sector(df_eval, "Luna")
        luna_bajo         = count_sector(df_eval, "Luna", "RIESGO BAJO")
        luna_medio        = count_sector(df_eval, "Luna", "RIESGO MEDIO")
        luna_alto         = count_sector(df_eval, "Luna", "RIESGO ALTO")
        luna_con_plan     = len(ids_con_plan_luna)
        luna_sin_bajo     = luna_bajo  - sum(1 for id_ in ids_con_plan_luna if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO BAJO")
        luna_sin_medio    = luna_medio - sum(1 for id_ in ids_con_plan_luna if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO MEDIO")
        luna_sin_alto     = luna_alto  - sum(1 for id_ in ids_con_plan_luna if not df_eval[df_eval["ID Evaluación"]==id_].empty and df_eval[df_eval["ID Evaluación"]==id_].iloc[0].get("Nivel","").strip().upper()=="RIESGO ALTO")
        luna_egreso       = count_bool_col(df_eval, "Luna", "egreso_alta") + count_bool_col(df_eval, "Luna", "egreso_traslado") + count_bool_col(df_eval, "Luna", "egreso_derivacion") + count_bool_col(df_eval, "Luna", "egreso_abandono")
        luna_egreso_alta  = count_bool_col(df_eval, "Luna", "egreso_alta")
        luna_egreso_tras  = count_bool_col(df_eval, "Luna", "egreso_traslado")
        luna_egreso_deriv = count_bool_col(df_eval, "Luna", "egreso_derivacion")
        luna_egreso_aban  = count_bool_col(df_eval, "Luna", "egreso_abandono")

        # Totales
        tot_inscritas  = n_inscritas_sol + n_inscritas_luna
        tot_eval       = sol_eval + luna_eval
        tot_bajo       = sol_bajo + luna_bajo
        tot_medio      = sol_medio + luna_medio
        tot_alto       = sol_alto + luna_alto
        tot_con_plan   = sol_con_plan + luna_con_plan
        tot_sin_bajo   = max(sol_sin_bajo + luna_sin_bajo, 0)
        tot_sin_medio  = max(sol_sin_medio + luna_sin_medio, 0)
        tot_sin_alto   = max(sol_sin_alto + luna_sin_alto, 0)
        tot_egreso     = sol_egreso + luna_egreso
        tot_eg_alta    = sol_egreso_alta + luna_egreso_alta
        tot_eg_tras    = sol_egreso_tras + luna_egreso_tras
        tot_eg_deriv   = sol_egreso_deriv + luna_egreso_deriv
        tot_eg_aban    = sol_egreso_aban + luna_egreso_aban

        # ------- Construir tabla REM-P7 -------
        # Estructura: lista de filas, cada fila = lista de celdas
        H_urbano = ["Clasificación de las familias por sector", "TOTAL", "Sector Sol\n(Urbano)", "Sector 2", "Sector 3", "Sector 4", "Sector 5", "Sector 6", "Sector 7", "Sector 8"]
        H_rural  = ["Clasificación de las familias por sector", "TOTAL", "Sector Luna\n(Rural)", "Sector 2", "Sector 3", "Sector 4", "Sector 5", "Sector 6", "Sector 7", "Sector 8"]
        H_inter  = ["Intervención en familias",                 "TOTAL", "Sol (Urbano)", "Luna\n(Rural)", "Sector 3", "Sector 4", "Sector 5", "Sector 6", "Sector 7", "Sector 8"]

        rows = [
            ["REM-P7. FAMILIAS EN CONTROL DE SALUD FAMILIAR"],
            [f"CESFAM Cholchol | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
            ["SECCIÓN A. CLASIFICACIÓN DE LAS FAMILIAS SECTOR URBANO (Sector Sol)"],
            H_urbano,
            ["N° Familias inscritas",                              n_inscritas_sol, n_inscritas_sol, "", "", "", "", "", "", ""],
            ["N° Familias evaluadas con cartola/encuesta familiar", sol_eval,      sol_eval,        "", "", "", "", "", "", ""],
            ["N° De familias en riesgo bajo",                       sol_bajo,      sol_bajo,        "", "", "", "", "", "", ""],
            ["N° De familias en riesgo medio",                      sol_medio,     sol_medio,       "", "", "", "", "", "", ""],
            ["N° De familias en riesgo alto",                       sol_alto,      sol_alto,        "", "", "", "", "", "", ""],
            ["SECCIÓN A.1 CLASIFICACIÓN DE LAS FAMILIAS SECTOR RURAL (Sector Luna)"],
            H_rural,
            ["N° Familias inscritas",                              n_inscritas_luna, n_inscritas_luna, "", "", "", "", "", "", ""],
            ["N° Familias evaluadas con cartola/encuesta familiar", luna_eval,      luna_eval,        "", "", "", "", "", "", ""],
            ["N° De familias en riesgo bajo",                       luna_bajo,      luna_bajo,        "", "", "", "", "", "", ""],
            ["N° De familias en riesgo medio",                      luna_medio,     luna_medio,       "", "", "", "", "", "", ""],
            ["N° De familias en riesgo alto",                       luna_alto,      luna_alto,        "", "", "", "", "", "", ""],
            ["SECCIÓN B. INTERVENCIÓN EN FAMILIAS SECTOR URBANO Y RURAL"],
            H_inter,
            ["N° Familias con plan de intervención",                tot_con_plan,  sol_con_plan,  luna_con_plan,  "", "", "", "", "", ""],
            ["N° Familias sin plan de intervención", "Riesgo bajo", tot_sin_bajo,   sol_sin_bajo,  luna_sin_bajo,  "", "", "", "", ""],
            ["",                                     "Riesgo medio",tot_sin_medio,  sol_sin_medio, luna_sin_medio, "", "", "", "", ""],
            ["",                                     "Riesgo alto", tot_sin_alto,   sol_sin_alto,  luna_sin_alto,  "", "", "", "", ""],
            ["N° Familias egresadas de planes de intervención", "Total de egresos", tot_egreso,     sol_egreso,    luna_egreso,    "", "", "", "", ""],
            ["",  "Alta por cumplir plan",                           tot_eg_alta,    sol_egreso_alta,  luna_egreso_alta,  "", "", "", "", ""],
            ["",  "Traslado de establecimiento",                     tot_eg_tras,    sol_egreso_tras,  luna_egreso_tras,  "", "", "", "", ""],
            ["",  "Derivación por complejidad",                      tot_eg_deriv,   sol_egreso_deriv, luna_egreso_deriv, "", "", "", "", ""],
            ["",  "Por abandono",                                    tot_eg_aban,    sol_egreso_aban,  luna_egreso_aban,  "", "", "", "", ""],
        ]

        # ------- Escribir en Hoja REM-P7 -------
        ws_rem = get_or_create_worksheet(spreadsheet, "REM-P7")
        ws_rem.clear()
        ws_rem.update(range_name="A1", values=rows)

        return True, f"Hoja REM-P7 actualizada ({tot_eval} evaluaciones procesadas)."

    except Exception as e:
        return False, f"Error actualizando REM-P7: {e}"


# --- ESTADO DE LA APLICACIÓN ---
if 'family_members' not in st.session_state:
    st.session_state.family_members = pd.DataFrame({
        "Nombre y Apellidos": pd.Series(dtype='str'),
        "RUT": pd.Series(dtype='str'),
        "F. Nac": pd.Series(dtype='object'),
        "Identidad de género": pd.Series(dtype='str'),
        "Pueblo Originario": pd.Series(dtype='str'),
        "Nacionalidad": pd.Series(dtype='str'),
        "E. Civil": pd.Series(dtype='str'),
        "Ocupación": pd.Series(dtype='str'),
        "Parentesco": pd.Series(dtype='str'),
        "Resp": pd.Series(dtype='bool')
    })
if 'intervention_plan' not in st.session_state:
    st.session_state.intervention_plan = pd.DataFrame({
        "Objetivo": pd.Series(dtype='str'),
        "Actividad": pd.Series(dtype='str'),
        "Fecha Prog": pd.Series(dtype='datetime64[ns]'),
        "Responsable": pd.Series(dtype='str'),
        "Fecha Real": pd.Series(dtype='datetime64[ns]'),
        "Evaluación": pd.Series(dtype='str'),
        "Estado": pd.Series(dtype='str'),
        "F. Seguimiento": pd.Series(dtype='datetime64[ns]'),
        "Obs. Seguimiento": pd.Series(dtype='str')
    })

if 'team_members' not in st.session_state:
    st.session_state.team_members = pd.DataFrame(columns=["Nombre y Profesión", "Firma"])
    
# Riesgos (Tablas 1-5)
risk_keys = [
    't1_vif', 't1_drogas', 't1_alcohol', 't1_saludMentalDescomp', 't1_abusoSexual', 
    't1_riesgoBiopsicoGrave', 't1_epsaRiesgo', 't1_vulnerabilidadExtrema', 't1_trabajoInfantil',
    't2_enfermedadGrave', 't2_altoRiesgoHosp', 't2_discapacidad', 't2_saludMentalLeve', 
    't2_judicial', 't2_rolesParentales', 't2_sobrecargaCuidador', 't2_conflictosSeveros', 't2_adultosRiesgo',
    't3_patologiaCronica', 't3_discapacidadLeve', 't3_rezago', 't3_madreAdolescente', 't3_duelo', 
    't3_sinRedApoyo', 't3_cesantia', 't3_vulneNoExtrema', 't3_precariedadLaboral', 
    't3_hacinamiento', 't3_entornoInseguro', 't3_adultoSolo', 't3_desercionEscolar', 
    't3_analfabetismo', 't3_escolaridadIncompleta', 't3_dificultadAcceso',
    't4_monoparental', 't4_riesgoCardio', 't4_contaminacion', 't4_higiene', 
    't4_sinRecreacion', 't4_sinEspaciosSeguros', 't4_endeudamiento', 't4_serviciosIncompletos',
    't5_lactancia', 't5_habitos', 't5_redesSociales', 't5_redFamiliar', 
    't5_comunicacion', 't5_recursosSuficientes', 't5_resiliencia', 't5_viviendaAdecuada'
]

for key in risk_keys:
    if key not in st.session_state:
        st.session_state[key] = False

def render_login_page():
    """Pantalla de login robusta y premium utilizando st.form como tarjeta."""
    st.markdown("""
        <style>
        /* Fondo Premium solo para main */
        [data-testid="stAppViewContainer"] > .main {
            background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%) !important;
        }
        
        /* Ocultar elementos sobrantes en login */
        header[data-testid="stHeader"] { background: transparent !important; }
        
        /* Estilizar la tarjeta del formulario nativo de Streamlit en el login */
        div[data-testid="stForm"] {
            background: rgba(255, 255, 255, 0.95) !important;
            backdrop-filter: blur(10px) !important;
            border-radius: 24px !important;
            padding: 40px !important;
            box-shadow: 0 10px 40px -10px rgba(0,0,0,0.1), 0 0 20px rgba(59, 130, 246, 0.05) !important;
            border: 1px solid rgba(255, 255, 255, 0.4) !important;
        }
        
        .login-title {
            color: #0f172a;
            font-size: 28px;
            font-weight: 800;
            text-align: center;
            margin-bottom: 8px;
            font-family: 'Inter', sans-serif;
            letter-spacing: -0.025em;
        }
        .login-subtitle {
            color: #64748b;
            font-size: 14px;
            text-align: center;
            margin-bottom: 24px;
            font-weight: 500;
        }
        .login-brand-wrapper {
            display: flex;
            justify-content: center;
            margin-bottom: 16px;
        }
        .login-brand {
            background: #eff6ff;
            color: #2563eb;
            padding: 6px 16px;
            border-radius: 9999px;
            font-size: 12px;
            font-weight: 700;
            letter-spacing: 0.1em;
            border: 1px solid #dbeafe;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<br><br><br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1.3, 1])
    with col2:
        with st.form("login_form", clear_on_submit=False, border=False):
            st.markdown("""
                <div class="login-brand-wrapper"><div class="login-brand">SISTEMA MAIS</div></div>
                <div class="login-title">Bienvenido/a</div>
                <div class="login-subtitle">Gestión de Riesgo Familiar · CESFAM Cholchol</div>
            """, unsafe_allow_html=True)
            
            user = st.text_input("Profesional", placeholder="Usuario registrado")
            password = st.text_input("Contraseña", type="password", placeholder="Su contraseña")
            st.markdown("<br>", unsafe_allow_html=True)
            submitted = st.form_submit_button("Ingresar a la Plataforma", width='stretch', type="primary")
            
            if submitted:
                users_df = load_users()
                if not users_df.empty:
                    user_match = users_df[users_df['usuario'].str.lower() == user.lower()]
                    if not user_match.empty:
                        actual_pass = str(user_match.iloc[0]['pass'])
                        if actual_pass == password:
                            st.session_state.authenticated = True
                            st.session_state.user_info = user_match.iloc[0].to_dict()
                            # Auditoría de Login
                            log_audit_event(st.session_state.user_info, "Inicio de Sesión", "Ingreso exitoso a la plataforma")
                            # Limpiar campos del formulario para garantizar ficha en blanco
                            for _k in ['idEvaluacion', 'familia', 'direccion', 'establecimiento',
                                       'sector', 'parentesco', 'programa_unidad', 'tipo_union',
                                       'evaluadorName', 'fechaEgreso']:
                                if _k in st.session_state:
                                    del st.session_state[_k]
                            st.rerun()
                        else:
                            st.error("Credenciales incorrectas.")
                    else:
                        st.error("Usuario no registrado.")
                else:
                    st.error("Error de base de datos.")

def main():
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
        
    if not st.session_state.authenticated:
        render_login_page()
        return

    # Si está autenticado, cargar sidebar con info de usuario
    user_info = st.session_state.user_info
    with st.sidebar:
        st.markdown(f"""
            <div style='padding: 10px; background: #f8fafc; border-radius: 5px; border-left: 5px solid #1F3864; margin-bottom: 20px;'>
                <div style='font-size: 12px; color: #64748b;'>Profesional:</div>
                <div style='font-weight: bold; color: #1e293b;'>{user_info.get('usuario', '').upper()}</div>
                <div style='font-size: 11px; color: #475569;'>{user_info.get('cargo', '')} | {user_info.get('Programa/Unidad', '')}</div>
            </div>
        """, unsafe_allow_html=True)
        
        if st.button("Cerrar Sesión", width='stretch'):
            # Limpieza atómica de toda la sesión para evitar fugas de datos RBAC
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    # --- MODO SIMULACIÓN (Solo Programador) ---
    # Guardamos el rol real en una variable para mantener el acceso al panel de simulación
    if 'real_role' not in st.session_state:
        st.session_state.real_role = str(st.session_state.user_info.get('rol', '')).lower()
    
    if st.session_state.real_role == 'programador':
        with st.sidebar:
            st.markdown("---")
            st.markdown("🛠️ **Zona de Pruebas (Simulación)**")
            sim_profile = st.selectbox(
                "Simular Perfil:",
                ["Original", "Jefe Sector Sol", "Jefe Sector Luna", "Encargado/a Postas", "Equipo Sector Sol", "Equipo Sector Luna"],
                index=0,
                key="sim_profile_selector"
            )
            
            if sim_profile != "Original":
                # Al simular, degradamos el 'rol' para que los filtros RBAC se activen
                st.session_state.user_info['rol'] = "usuario" 
                
                if sim_profile == "Jefe Sector Sol":
                    st.session_state.user_info['cargo'] = "Jefe Sector Sol"
                    st.session_state.user_info['Programa/Unidad'] = "Sector Sol"
                elif sim_profile == "Jefe Sector Luna":
                    st.session_state.user_info['cargo'] = "Jefe Sector Luna"
                    st.session_state.user_info['Programa/Unidad'] = "Sector Luna"
                elif sim_profile == "Encargado/a Postas":
                    st.session_state.user_info['cargo'] = "Encargado/a Postas"
                    st.session_state.user_info['Programa/Unidad'] = "Postas Salud Rural"
                elif sim_profile == "Equipo Sector Sol":
                    st.session_state.user_info['cargo'] = "Equipo de Sector"
                    st.session_state.user_info['Programa/Unidad'] = "Equipo de Sector Sol"
                elif sim_profile == "Equipo Sector Luna":
                    st.session_state.user_info['cargo'] = "Equipo de Sector"
                    st.session_state.user_info['Programa/Unidad'] = "Equipo de Sector Luna"
                
                st.warning(f"Simulando: **{sim_profile}** (Privilegios restringidos)")
                # Forzar recarga de datos con el nuevo filtro RBAC
                if 'df_evaluaciones' in st.session_state:
                    del st.session_state['df_evaluaciones']
                if 'raw_analytics_df' in st.session_state:
                    del st.session_state['raw_analytics_df']
            else:
                # Restaurar rol original
                st.session_state.user_info['rol'] = st.session_state.real_role
                # Restaurar cargo/unidad si es necesario (el login original está en st.session_state.user_info)
                # Nota: El login original se perdió al sobrescribir, pero para el Programador
                # podemos resetearlo a valores genéricos o recargar.
                if st.session_state.user_info['rol'] == 'programador':
                    st.session_state.user_info['cargo'] = "Programador"
                    st.session_state.user_info['Programa/Unidad'] = "Administración"
            
    # El resto del main sigue aquí... (listado, pestañas, etc.)
    
    # Init Session State variables for Inputs if they don't exist
    input_keys = [
        'idEvaluacion', 'fechaEvaluacion', 'familia', 'direccion',
        'establecimiento', 'sector', 'evaluadorName', 'parentesco', 'programa_unidad',
        'tipo_union'
    ]
    for k in input_keys:
        if k not in st.session_state:
            st.session_state[k] = ""
            if k == 'fechaEvaluacion': st.session_state[k] = date.today()
            if k == 'sector': st.session_state[k] = "Sol"
            if k == 'parentesco': st.session_state[k] = PARENTESCO_OPTIONS[0]
            if k == 'programa_unidad': st.session_state[k] = PROGRAMA_OPTIONS[0]
            if k == 'tipo_union': st.session_state[k] = TIPO_UNION_OPTIONS[0]


    # --- SIDEBAR GESTIÓN ---
    with st.sidebar:
        st.markdown('<div style="color: #0ea5e9; font-size: 0.75rem; font-weight: 800; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 4px; margin-top: -30px;">SISTEMA MAIS</div>', unsafe_allow_html=True)
        st.markdown('<div style="color: #0f172a; font-size: 1.25rem; font-weight: 800; letter-spacing: -0.02em; margin-bottom: 16px;">🏥 Panel Clínico</div>', unsafe_allow_html=True)
        st.markdown('<div style="color: #475569; font-size: 0.85rem; line-height: 1.5; margin-bottom: 20px;">Gestione y busque los registros consolidados en la base médica centralizada.</div>', unsafe_allow_html=True)
        
        with st.container(border=True):
            if st.button("➕ Nueva Ficha", width='stretch'):
                keys_to_clear = [
                    'idEvaluacion', 'familia', 'direccion', 'establecimiento',
                    'sector', 'parentesco', 'programa_unidad', 'tipo_union',
                    'evaluadorName', 'fechaEgreso', 'comp_rep_sector', 'comp_familia',
                    'comp_dir', 'comp_rep_fam', 'comp_rut', 'comp_fecha',
                    'sig_func', 'sig_benef', 'sig_equipo', 'sig_jefe', 'sig_evaluador_input',
                    'egreso_alta', 'egreso_traslado', 'egreso_derivacion', 'egreso_abandono',
                    'observaciones'
                ]
                for _k in keys_to_clear:
                    if _k in st.session_state:
                        del st.session_state[_k]
                for rk in risk_keys:
                    st.session_state[rk] = False
                st.session_state.family_members = pd.DataFrame({
                    "Nombre y Apellidos": pd.Series(dtype='str'),
                    "RUT": pd.Series(dtype='str'),
                    "F. Nac": pd.Series(dtype='object'),
                    "Identidad de género": pd.Series(dtype='str'),
                    "Pueblo Originario": pd.Series(dtype='str'),
                    "Nacionalidad": pd.Series(dtype='str'),
                    "E. Civil": pd.Series(dtype='str'),
                    "Ocupación": pd.Series(dtype='str'),
                    "Parentesco": pd.Series(dtype='str'),
                    "Resp": pd.Series(dtype='bool')
                })
                st.session_state.intervention_plan = pd.DataFrame({
                    "Objetivo": pd.Series(dtype='str'), "Actividad": pd.Series(dtype='str'),
                    "Fecha Prog": pd.Series(dtype='datetime64[ns]'), "Responsable": pd.Series(dtype='str'),
                    "Fecha Real": pd.Series(dtype='datetime64[ns]'), "Evaluación": pd.Series(dtype='str'),
                    "Estado": pd.Series(dtype='str'), "F. Seguimiento": pd.Series(dtype='datetime64[ns]'),
                    "Obs. Seguimiento": pd.Series(dtype='str')
                })
                st.session_state.team_members = pd.DataFrame(columns=["Nombre y Profesión", "Firma"])
                st.rerun()

        with st.container(border=True):
            st.markdown('<div style="font-weight: 700; font-size: 0.9rem; color: #334155; margin-bottom: 8px;">Búsqueda Directa</div>', unsafe_allow_html=True)
            search_id = st.text_input("ID Evaluación", placeholder="Ej: FAM-0123...", label_visibility="collapsed")
            if st.button("🔍 Cargar Registro", type="primary", width='stretch'):
                with st.spinner("Conectando con base de datos segura..."):
                    record = search_record(search_id)
                    if record:
                        load_record_into_state(record)
                        
                        st.session_state['idEvaluacion'] = str(record.get('ID Evaluación', ''))
                        st.session_state['familia'] = record.get('Familia', '')
                        st.session_state['direccion'] = record.get('Dirección', '')
                        st.session_state['establecimiento'] = record.get('Establecimiento', '')
                        
                        raw_sector = record.get('Sector', 'Sol')
                        if raw_sector.upper() == 'LUNA':
                            cleaned_sector = 'Luna'
                        elif raw_sector in ['No Asignado', 'NO_ESPECIFICADO']:
                            cleaned_sector = 'No identificado'
                        else:
                            cleaned_sector = raw_sector
                        st.session_state['sector'] = cleaned_sector

                        # Nuevos campos
                        raw_parentesco = record.get('Parentesco', PARENTESCO_OPTIONS[0])
                        st.session_state['parentesco'] = raw_parentesco if raw_parentesco in PARENTESCO_OPTIONS else PARENTESCO_OPTIONS[0]
                        
                        raw_programa = record.get('Programa/Unidad', PROGRAMA_OPTIONS[0])
                        st.session_state['programa_unidad'] = raw_programa if raw_programa in PROGRAMA_OPTIONS else PROGRAMA_OPTIONS[0]
                        
                        st.success(f"✅ Registro {search_id} cargado.")
                        st.rerun()
                    else:
                        st.error("❌ El ID ingresado no coincide con ningún registro.")
        
        # --- NUEVO: DESCARGA FORMULARIO BLANCO ---
        st.markdown('<div style="font-weight: 700; font-size: 0.9rem; color: #334155; margin-top: 10px; margin-bottom: 8px;">Instrumentos en Blanco</div>', unsafe_allow_html=True)
        with st.container(border=True):
            st.markdown('<div style="font-size: 0.8rem; color: #64748b; margin-bottom: 10px;">Descargue la pauta vacía para aplicación manual en terreno.</div>', unsafe_allow_html=True)
            
            def get_blank_pdf_bytes():
                return generate_blank_pdf()
            
            blank_pdf = get_blank_pdf_bytes()
            st.download_button(
                label="📄 Descargar Pauta en Blanco",
                data=blank_pdf,
                file_name=f"Pauta_Blank_{date.today()}_{uuid.uuid4().hex[:6]}.pdf",
                mime="application/pdf",
                width='stretch'
            )

        col_busq, col_est = st.columns([2, 1])
        with col_busq:
            search_query = st.text_input("🔍 Buscar por Familia, RUT o Dirección:", placeholder="Ej: Perez, 12.345.678-9...", key="search_main")
        with col_est:
            # Opciones de establecimiento para filtro
            est_options = ["Todos", "Cesfam Cholchol", "Posta Huentelar", "Posta Huamaqui", "Posta Malalche", "EMR Rapahue", "EMR Repocura"]
            
            # Pre-selección inteligente por rol
            default_est_idx = 0
            u_cargo = str(st.session_state.user_info.get('cargo', '')).lower()
            if 'encargado' in u_cargo and 'postas' in u_cargo:
                # Si es de postas, por defecto no mostramos "Todos" (Cesfam incluído)
                # sino que invitamos a filtrar o mostramos el primero de postas
                # O simplemente dejamos "Todos" pero el backend ya filtró por Sector Luna
                pass
                
            selected_est_filter = st.selectbox("Filtrar por Establecimiento:", options=est_options, index=default_est_idx, key="filter_est_main")

        # Cargar datos para el listado y búsqueda (Usamos la función centralizada de analytics)
        from analytics import load_evaluaciones_df
        df_display = load_evaluaciones_df(est_filter=selected_est_filter)
        
        # Filtro de búsqueda (se aplica sobre los datos ya filtrados por RBAC y Establecimiento)
        if search_query:
            q = search_query.lower()
            mask = (
                df_display['Familia'].astype(str).str.lower().str.contains(q, na=False) | 
                df_display['Dirección'].astype(str).str.lower().str.contains(q, na=False) |
                df_display['Grupo Familiar JSON'].astype(str).str.lower().str.contains(q, na=False)
            )
            df_display = df_display[mask]

        with st.expander("📋 Mis Encuestas Familiares"):
            st.caption(f"Fichas autorizadas para: **{selected_est_filter}**")
            # En este punto df_display YA está filtrado por RBAC, Establecimiento y Búsqueda
            if not df_display.empty:
                # Mostrar columnas clave
                cols_show = ["ID Evaluación", "Familia", "Sector", "Nivel", "Establecimiento"]
                cols_exist = [c for c in cols_show if c in df_display.columns]
                st.dataframe(df_display[cols_exist], width='stretch', hide_index=True)
                st.info("💡 Copie el ID y búsquelo arriba para cargar los datos.")
            else:
                st.write("No se encontraron registros para los filtros seleccionados.")

        if can_download_rem(user_info):
            st.markdown("---")
            st.subheader("📊 Config. REM-P7")
            st.markdown("*Familias inscritas por sector (para el informe REM-P7):*")
            n_inscritas_sol  = st.number_input("Inscritas Sector Sol",  min_value=0, value=st.session_state.get('n_inscritas_sol', 0),  step=1, key="n_inscritas_sol")
            n_inscritas_luna = st.number_input("Inscritas Sector Luna", min_value=0, value=st.session_state.get('n_inscritas_luna', 0), step=1, key="n_inscritas_luna")
            if st.button("🔄 Actualizar REM-P7 (Sheets)", width='stretch'):
                with st.spinner("Generando REM-P7..."):
                    ok, msg = update_rem_p7(n_inscritas_sol, n_inscritas_luna)
                    if ok:
                        # Auditoría REM-P7
                        log_audit_event(st.session_state.user_info, "Generación REM-P7", f"Reporte REM-P7 actualizado en Sheets")
                        st.success(f"✅ {msg}")
                    else:
                        st.error(f"❌ {msg}")

            st.markdown("**Exportar Excel:**")
            if st.button("📥 Descargar REM-P7 Excel", width='stretch'):
                with st.spinner("Generando Excel..."):
                    buf, err = export_rem_p7_excel(n_inscritas_sol, n_inscritas_luna)
                    if err:
                        st.error(f"❌ {err}")
                    else:
                        st.session_state['rem_p7_excel'] = buf
                        st.success("✅ Excel listo para descargar.")

            if st.session_state.get('rem_p7_excel'):
                fname = f"REM-P7_{date.today().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="⬇️ Guardar archivo Excel",
                    data=st.session_state['rem_p7_excel'],
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch'
                )
        else:
            st.info("ℹ️ Su cargo no tiene permisos para descargar el REM-P7.")

        if str(user_info.get('rol', '')).lower() == 'programador':
            st.markdown("---")
            st.subheader("🔧 Administración de Datos")
            with st.expander("🔄 Migrar IDs al nuevo formato"):
                st.caption("Reescribe todos los IDs de evaluación existentes al formato **EVA-NNN-FAM-XXX**.")
                st.warning("⚠️ Esta operación modifica los IDs de todos los registros en Google Sheets. Ejecutar solo una vez.")
                if st.button("🚀 Ejecutar Migración de IDs", type="primary", width='stretch'):
                    with st.spinner("Migrando IDs... puede tomar unos segundos por cada registro..."):
                        ok_m, msg_m, n_m = migrate_eval_ids_to_new_format()
                        if ok_m:
                            # Auditoría Migración
                            log_audit_event(st.session_state.user_info, "Migración de IDs", f"Se migraron {n_m} registros al nuevo formato")
                            st.success(f"✅ {msg_m}")
                            st.balloons()
                        else:
                            st.error(f"❌ {msg_m}")



    # --- SELECTOR DE VISTA ---
    vista = st.radio(
        "Vista:",
        ["📋 Ficha Familiar", "📊 Dashboard Analítico", "🔍 Análisis Familiar"],
        horizontal=True,
        label_visibility="collapsed",
        key="vista_selector"
    )

    # --- PROCESAMIENTO DE DATOS (Común a todas las vistas para evitar inconsistencias) ---
    # Esto asegura que st.session_state.family_members esté siempre al día
    # aunque estemos en la pestaña de análisis.
    
    # Nota: El data_editor solo se muestra en la Ficha Familiar, pero manejamos
    # su estado globalmente si es posible. 
    # Sin embargo, Streamlit solo actualiza el valor del widget cuando se renderiza.
    # Por lo tanto, si estamos en Análsis, usamos lo que ya esté en session_state.

    if vista == "📊 Dashboard Analítico":
        if render_analytics:
            # FILTRO RBAC: Solo pasar los datos que el usuario puede ver
            df_full = load_evaluaciones_df() # Esta función está en analytics.py pero app.py carga sus propios datos
            # En app.py no importamos load_evaluaciones_df, así que usaremos la lógica local o pediremos a analytics que filtre
            
            # Nota: analytics.py tiene su propia función de carga. Vamos a modificar analytics.py para que respete el session_state
            render_analytics()
        else:
            st.warning("⚠️ plotly no instalado. Reinicia la app después de instalar.")
        st.stop()

    if vista == "🔍 Análisis Familiar":
        # SEGURIDAD RBAC: Validar que el registro cargado sea accesible para el usuario
        current_id = st.session_state.get('idEvaluacion', '')
        if current_id:
            # Creamos un mock del registro para validar acceso
            mock_row = {
                'ID Evaluación': current_id,
                'Sector': st.session_state.get('sector', ''),
                'Programa/Unidad': st.session_state.get('programa_unidad', '')
            }
            if not check_access(mock_row, user_info):
                st.error("🚫 ACCESO DENEGADO: No tiene permisos para visualizar este estudio de familia.")
                st.info("Solo los responsables del sector o programa respectivo pueden acceder a este análisis clínico.")
                st.stop()
                
        # 1. DATOS BASE
        members_list = st.session_state.family_members.to_dict(orient='records')
        familia_val  = st.session_state.get('familia', 'Sin Nombre')
        prog_val = st.session_state.get('programa_unidad', 'APS General')
        eval_id = st.session_state.get('idEvaluacion', 'N/A')
        
        # 2. CÁLCULO DE RIESGOS (Protocolo San Juan: 10, 4, 2 pts)
        risk_keys_list = [
            't1_vif','t1_drogas','t1_alcohol','t1_saludMentalDescomp','t1_abusoSexual',
            't1_riesgoBiopsicoGrave','t1_epsaRiesgo','t1_vulnerabilidadExtrema','t1_trabajoInfantil',
            't2_enfermedadGrave','t2_altoRiesgoHosp','t2_discapacidad','t2_saludMentalLeve',
            't2_judicial','t2_rolesParentales','t2_adultosRiesgo',
            't3_patologiaCronica','t3_discapacidadLeve','t3_rezago','t3_madreAdolescente',
            't3_sinRedApoyo','t3_cesantia','t3_vulneNoExtrema','t3_precariedadLaboral',
            't3_hacinamiento','t3_entornoInseguro','t3_adultoSolo','t3_desercionEscolar',
            't3_analfabetismo','t3_escolaridadIncompleta','t3_dificultadAcceso',
            't4_monoparental','t4_riesgoCardio','t4_contaminacion','t4_higiene',
            't4_sinRecreacion','t4_sinEspaciosSeguros','t4_endeudamiento','t4_serviciosIncompletos',
            't5_lactancia','t5_habitos','t5_redesSociales','t5_redFamiliar',
            't5_comunicacion','t5_recursosSuficientes','t5_resiliencia','t5_viviendaAdecuada'
        ]
        active_risks = {k: bool(st.session_state.get(k, False)) for k in risk_keys_list}
        
        _t1 = sum(1 for k in ['t1_vif','t1_drogas','t1_alcohol','t1_saludMentalDescomp'] if active_risks.get(k))
        _t2 = sum(1 for k in ['t2_enfermedadGrave','t2_altoRiesgoHosp','t2_discapacidad'] if active_risks.get(k))
        _pts = sum(4 for k in risk_keys_list if k.startswith('t3_') and active_risks.get(k)) + \
               sum(3 for k in risk_keys_list if k.startswith('t4_') and active_risks.get(k))
        
        nivel_val = "RIESGO ALTO" if (_t1 >= 1 or _t2 >= 2 or _pts >= 26) else \
                    "RIESGO MEDIO" if (_t2 == 1 or 17 <= _pts <= 25) else "RIESGO BAJO"

        # 3. RENDERIZADO DE INTERFAZ PREMIUM (HTML/CSS)
        active_list = [(k[:2].upper(), RISK_LABELS.get(k, k)) for k, v in active_risks.items() if v and not k.startswith('t5_')] # Omite protectores
        
        # Colores semánticos
        badge_color = "#10b981" if nivel_val == "RIESGO BAJO" else "#f59e0b" if nivel_val == "RIESGO MEDIO" else "#ef4444"
        badge_bg = "#d1fae5" if nivel_val == "RIESGO BAJO" else "#fef3c7" if nivel_val == "RIESGO MEDIO" else "#fee2e2"
        badge_text = "#064e3b" if nivel_val == "RIESGO BAJO" else "#78350f" if nivel_val == "RIESGO MEDIO" else "#7f1d1d"

        # HTML para pills de riesgo
        pills_html = ""
        for tipo, label in active_list[:5]:
            pills_html += f'<span style="display:inline-block; padding: 4px 10px; margin: 4px; background: #f3f4f6; color: #4b5563; border-radius: 9999px; font-size: 0.75rem; font-weight: 500; border: 1px solid #e5e7eb;">{tipo}: {label}</span>'
        if len(active_list) > 5:
            pills_html += f'<span style="display:inline-block; padding: 4px 10px; margin: 4px; background: #f3f4f6; color: #4b5563; border-radius: 9999px; font-size: 0.75rem; font-weight: 500; border: 1px solid #e5e7eb;">+{len(active_list)-5} más...</span>'
        if not pills_html:
            pills_html = '<span style="color: #10b981; font-size: 0.85rem; font-weight: 500;">✅ Sin riesgos activos detectados</span>'

        date_str = date.today().strftime('%d/%b/%Y')

        # Tarjeta "Clean Medical"
        card_html = f"""
<div style="background: white; border-radius: 16px; padding: 24px; box-shadow: 0 4px 20px rgba(0,0,0,0.05); margin-bottom: 24px; border: 1px solid #f3f4f6; font-family: 'Inter', sans-serif;">
    <div style="display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px;">
        <div>
            <h2 style="margin: 0; font-size: 1.5rem; color: #111827; font-weight: 700; letter-spacing: -0.025em;">Estudio de Familia: {familia_val}</h2>
            <p style="margin: 4px 0 0 0; color: #6b7280; font-size: 0.875rem;">ID: {eval_id} &nbsp;•&nbsp; Programa: {prog_val}</p>
        </div>
        <div style="background: {badge_bg}; color: {badge_text}; padding: 6px 14px; border-radius: 9999px; font-weight: 700; font-size: 0.85rem; border: 1px solid {badge_color}40; letter-spacing: 0.025em;">
            {nivel_val}
        </div>
    </div>
    
    <div style="display: flex; gap: 16px; margin-bottom: 24px; overflow-x: auto;">
        <div style="flex: 1; min-width: 120px; background: #f8fafc; padding: 16px; border-radius: 12px; border: 1px solid #e2e8f0; text-align: center;">
            <div style="color: #64748b; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">Integrantes</div>
            <div style="color: #0f172a; font-size: 1.75rem; font-weight: 800; margin-top: 4px;">{len(members_list)}</div>
        </div>
        <div style="flex: 1; min-width: 120px; background: #f8fafc; padding: 16px; border-radius: 12px; border: 1px solid #e2e8f0; text-align: center;">
            <div style="color: #64748b; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">Puntaje Total</div>
            <div style="color: #0f172a; font-size: 1.75rem; font-weight: 800; margin-top: 4px;">{_pts} <span style="font-size: 1rem; color: #64748b; font-weight: 600;">pts</span></div>
        </div>
        <div style="flex: 1; min-width: 120px; background: #f8fafc; padding: 16px; border-radius: 12px; border: 1px solid #e2e8f0; text-align: center;">
            <div style="color: #64748b; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em;">Factores Activos</div>
            <div style="color: #0f172a; font-size: 1.75rem; font-weight: 800; margin-top: 4px; color: {badge_color};">{len(active_list)}</div>
        </div>
    </div>

    <div style="background: #ffffff; padding-top: 8px;">
        <div style="color: #374151; font-size: 0.875rem; font-weight: 700; margin-bottom: 8px;">Riesgos Principales:</div>
        <div style="display: flex; flex-wrap: wrap; gap: 4px; margin-left: -4px;">
            {pills_html}
        </div>
    </div>
</div>
"""
        st.html(card_html)

        # --- LAYOUT DE ESTUDIO ---

        tab_graficos, tab_riesgos, tab_plan = st.tabs(["📊 Gráficos Clínicos", "⚠️ Detalle de Riesgos", "📋 Plan de Intervención"])

        with tab_graficos:
            # Controles para el Ecomapa (Construcción manual)
            st.markdown("##### 🛠️ Configuración de Redes (Ecomapa)")
            all_systems = ["CESFAM", "RELIGIÓN", "TRABAJO", "ESCUELA", "COMUNIDAD", "JUSTICIA", "RED FAMILIAR", "VECINOS", "OTRO"]
            col_sel1, col_sel2 = st.columns(2)
            with col_sel1:
                selected_systems = st.multiselect("Sistemas vinculados:", all_systems, default=["CESFAM", "COMUNIDAD"])
            with col_sel2:
                # Selector de flujos (Energy Flow SJ Protocol)
                with st.expander("🏹 Sentido del Vínculo (Flujos)"):
                    system_flows = {}
                    for sys in selected_systems:
                        flow = st.radio(f"Dirección para {sys}:", 
                                       ["Recíproco ↔", "Hacia Familia ←", "Desde Familia →", "Sin flujo"],
                                       horizontal=True, key=f"flow_{sys}")
                        
                        # Mapeo a términos técnicos de ecomap.py
                        mapping = {"Recíproco ↔": "both", "Hacia Familia ←": "in", "Desde Familia →": "out", "Sin flujo": "none"}
                        system_flows[sys] = mapping[flow]

            st.markdown("---")
            col_gen, col_eco = st.columns(2)
            
            with col_gen:
                st.markdown("#### 🌳 Genograma Estructurado")
                if generate_genogram_dot and members_list:
                    # Pasar el tipo de unión seleccionado
                    dot_geno = generate_genogram_dot(
                        members_list, 
                        familia_val, 
                        nivel_val, 
                        tipo_union=st.session_state.get('tipo_union', 'Casados'),
                        interpersonal_relations=st.session_state.get('interpersonal_relations', [])
                    )
                    st.graphviz_chart(dot_geno, width='stretch')
                else:
                    st.info("💡 Agregue integrantes en la pestaña 'Ficha Familiar'.")
                    
            with col_eco:
                st.markdown("#### 🕸️ Ecomapa de Redes")
                if generate_ecomap_dot:
                    # Pasamos los sistemas seleccionados y sus flujos
                    dot_eco = generate_ecomap_dot(familia_val, members_list, active_risks, prog_val, nivel_val, selected_systems, system_flows)
                    st.graphviz_chart(dot_eco, width='stretch')
                else:
                    st.error("Error modular en Ecomapa.")
            
            # Narrativa Clínica
            st.markdown("---")
            st.markdown(generate_clinical_narrative(members_list, active_risks, nivel_val, prog_val))

        with tab_riesgos:
            st.markdown("#### 🔍 Detalle de Factores de Riesgo Detectados")
            # Usar RISK_LABELS para mostrar nombres clínicos reales
            active_list = [(k[:2].upper(), RISK_LABELS.get(k, k)) for k, v in active_risks.items() if v]
            
            if active_list:
                cols = st.columns(2)
                for i, (tipo, label) in enumerate(active_list):
                    cols[i % 2].warning(f"🔸 **({tipo})** {label}")
            else:
                st.success("✅ No se detectan factores de riesgo activos.")

        with tab_plan:
            st.markdown("#### 📅 Plan de Intervención Familiar Actual")
            if not st.session_state.intervention_plan.empty:
                st.dataframe(st.session_state.intervention_plan, width='stretch', hide_index=True)
            else:
                st.info("💡 No se ha definido un plan de intervención para esta familia.")

        st.markdown("---")
        # Barra de acciones de estudio
        col_s1, col_s2 = st.columns([1, 4])
        with col_s1:
            if st.button("💾 Guardar Estudio Completo", type="primary", width='stretch'):
                with st.spinner("Persistiendo estudio en el historial..."):
                    # 1. Preparar datos para Hoja 1 (Evaluaciones)
                    # ---- EXTRAER RUTs DEL GRUPO FAMILIAR ----
                    def normalizar_rut(rut_str):
                        return rut_str.replace(".", "").strip()

                    df_fam_rut = st.session_state.family_members.fillna("")
                    if 'RUT' in df_fam_rut.columns:
                        ruts_list = [normalizar_rut(str(r)) for r in df_fam_rut['RUT'].tolist() if str(r).strip()]
                        ruts_concatenados = ",".join(ruts_list)
                    else:
                        ruts_concatenados = ""

                    evaluador_n = st.session_state.get('evaluadorName', 'N/A')
                    data_row = [
                        eval_id, str(st.session_state.get('fechaEvaluacion', date.today())),
                        familia_val, st.session_state.get('direccion', ''), 
                        st.session_state.get('establecimiento', ''), st.session_state.get('sector', ''),
                        st.session_state.get('parentesco', ''), prog_val,
                        st.session_state.get('total_points', 0), nivel_val, evaluador_n,
                        st.session_state.get('tipo_union', 'Casados'), ruts_concatenados
                    ]
                    
                    # Riesgos
                    risk_list = [st.session_state.get(k, False) for k in risk_keys]
                    data_row.extend(risk_list)

                    # JSONs
                    df_fam = st.session_state.family_members.fillna("")
                    data_row.append(json.dumps(df_fam.to_dict('records'), ensure_ascii=False, default=str))
                    
                    df_plan_save = st.session_state.intervention_plan.copy()
                    for c in ['Fecha Prog', 'Fecha Real']:
                        if c in df_plan_save.columns:
                            df_plan_save[c] = df_plan_save[c].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notnull(x) and hasattr(x, 'strftime') else "")
                    data_row.append(json.dumps(df_plan_save.to_dict('records'), ensure_ascii=False))
                    
                    df_team = st.session_state.team_members.fillna("")
                    data_row.append(json.dumps(df_team.to_dict('records'), ensure_ascii=False, default=str))

                    rel_json = json.dumps(st.session_state.get('interpersonal_relations', []), ensure_ascii=False)
                    data_row.append(rel_json)
                    
                    # Seguimiento del Plan
                    df_seg_save = st.session_state.get('seguimiento_plan', pd.DataFrame()).copy()
                    data_row.append(json.dumps(df_seg_save.fillna('').to_dict('records'), ensure_ascii=False, default=str))
                    
                    # Extra data (APGAR)
                    data_row.extend([
                        st.session_state.get('apgar_total', 0),
                        st.session_state.get('apgar_a1', 0),
                        st.session_state.get('apgar_a2', 0),
                        st.session_state.get('apgar_a3', 0),
                        st.session_state.get('apgar_a4', 0),
                        st.session_state.get('apgar_a5', 0)
                    ])

                    # Extra data
                    extra_data = [
                        st.session_state.get('comp_rep_sector', ''), st.session_state.get('comp_familia', ''),
                        st.session_state.get('comp_dir', ''), st.session_state.get('comp_rep_fam', ''),
                        st.session_state.get('comp_rut', ''), str(st.session_state.get('comp_fecha', date.today())),
                        st.session_state.get('sig_func', ''), st.session_state.get('sig_benef', ''),
                        st.session_state.get('sig_equipo', ''), st.session_state.get('sig_jefe', ''),
                        st.session_state.get('sig_evaluador_input', ''),
                        st.session_state.get('egreso_alta', False), st.session_state.get('egreso_traslado', False),
                        st.session_state.get('egreso_derivacion', False), st.session_state.get('egreso_abandono', False),
                        str(st.session_state.get('fechaEgreso', '')) if st.session_state.get('fechaEgreso') else "",
                        st.session_state.get('observaciones', ''),
                        st.session_state.get('link_drive', '')
                    ]
                    data_row.extend(extra_data)

                    final_headers = [
                        "ID Evaluación", "Fecha", "Familia", "Dirección", "Establecimiento", "Sector",
                        "Parentesco", "Programa/Unidad",
                        "Puntaje", "Nivel", "Evaluador",
                        "Tipo Unión", "RUTs Grupo Familiar",
                    ] + risk_keys + [
                        "Grupo Familiar JSON", "Plan Intervención JSON", "Equipo Salud JSON", "Relaciones JSON", "Seguimiento Plan JSON",
                        "APGAR Total", "A1", "A2", "A3", "A4", "A5"
                    ] + [
                        "Rep Sector", "Familia Comp", "Dir Comp", "Rep Familia", "RUT Rep", "Fecha Comp",
                        "Firma Funcionario", "Firma Beneficiario", "Firma Equipo", "Firma Jefe", "Firma Evaluador",
                        "egreso_alta", "egreso_traslado", "egreso_derivacion", "egreso_abandono",
                        "Fecha Egreso", "Observaciones", "Carpeta Digital (Drive)"
                    ]

                    # 2. Guardar en Sheets
                    ok1, msg1 = save_evaluacion_to_sheet(data_row, final_headers)
                    ok2, msg2 = save_intervention_rows(eval_id, familia_val, str(st.session_state.get('fechaEvaluacion', date.today())), nivel_val, prog_val, st.session_state.get('parentesco', ''), st.session_state.intervention_plan)
                    
                    # 3. Guardar Ecomapa
                    ecomap_elements = {
                        "selected_systems": selected_systems, 
                        "system_flows": system_flows,
                        "active_risks": active_risks
                    }
                    ok3, msg3 = save_ecomap_to_sheet(eval_id, familia_val, ecomap_elements)
                    
                    if ok1 and ok2:
                        # Auditoría de Guardado/Actualización
                        accion_audit = "Actualización de Registro" if _es_registro_existente else "Creación de Registro"
                        log_audit_event(st.session_state.user_info, accion_audit, f"Evaluación guardada en Sheets. Familia: {familia_val}", eval_id=id_evaluacion)
                        
                        st.success(f"✅ Estudio Completo Guardado: {msg1} | {msg2} | {msg3}")
                        st.balloons()
                    else:
                        st.error(f"❌ Error al guardar: {msg1} {msg2}")
        with col_s2:
            st.info(f"Estudio consolidado de la Familia **{familia_val}** | ID: {eval_id}")

        st.stop()

    st.markdown("<br>", unsafe_allow_html=True)

    # --- 1. DATOS DE IDENTIFICACIÓN ---
    st.markdown('<div style="color: #0f172a; font-size: 1.1rem; font-weight: 700; margin-bottom: 8px;">1. Identificación y Contexto</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col_logo, col_title, col_id = st.columns([1.5, 3.5, 2])
        
        with col_logo:
            # Normalizado para compatibilidad con Windows local y Streamlit Cloud (Linux)
            LOGO_PATH = "NUEVO LOGO.png"
            if os.path.exists(LOGO_PATH):
                st.image(LOGO_PATH, width=120)
            else:
                st.markdown('<div style="border:1px solid #e2e8f0; padding:10px; border-radius:12px; width:80px; text-align:center; color:#64748b; background:#f8fafc;">LOGO</div>', unsafe_allow_html=True)
                
        with col_title:
            st.markdown("""
            <div style="padding-top: 10px;">
                <h2 style="margin:0; font-size:1.4rem; text-transform:uppercase; letter-spacing: -0.01em;">Ilustre Municipalidad de Cholchol</h2>
                <h4 style="margin:0; font-size:1rem; color:#64748b; text-transform:uppercase;">Departamento de Salud | CESFAM Cholchol</h4>
            </div>
            """, unsafe_allow_html=True)
            
        with col_id:
            st.markdown("**FICHA DE INGRESO**")
            id_actual = st.session_state.get('idEvaluacion', '')
            if id_actual:
                # Registro cargado desde el buscador
                st.markdown(
                    f"**N° ID Evaluación**\n\n"
                    f'<div style="background:#f0f9ff; border:1px solid #bae6fd; border-radius:10px; padding:8px 14px; font-size:1rem; font-weight:700; color:#0369a1; letter-spacing:0.05em;">'
                    f'🏷️ {id_actual}</div>',
                    unsafe_allow_html=True
                )
            else:
                # Registro nuevo: sin ID aún
                st.markdown(
                    "**N° ID Evaluación**\n\n"
                    '<div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:8px 14px; font-size:0.85rem; color:#94a3b8; font-style:italic;">'
                    '🔖 Se asignará al guardar</div>',
                    unsafe_allow_html=True
                )
            fecha_input = st.date_input("Fecha", key="fechaEvaluacion")


        st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 16px 0;'>", unsafe_allow_html=True)

        # 2. DATOS DE IDENTIFICACIÓN
        c1, c2 = st.columns(2)
        with c1:
            familia = st.text_input("Familia (Apellidos):", placeholder="Ej: Pérez González", key="familia")
            direccion = st.text_input("Dirección Completa:", key="direccion")
        with c2:
            establecimiento = st.selectbox(
                "Establecimiento Base:", 
                ["Cesfam Cholchol", "Posta Huentelar", "Posta Huamaqui", "Posta Malalche", "EMR Rapahue", "EMR Repocura"],
                key="establecimiento"
            )
            sector = st.selectbox("Sector Asignado:", ["Sol", "Luna", "No identificado"], key="sector")

        # Nuevos campos: Parentesco y Programa/Unidad
        c3, c4 = st.columns(2)
        with c3:
            parentesco = st.selectbox(
                "Parentesco del Evaluado (Informante):",
                PARENTESCO_OPTIONS,
                key="parentesco",
                help="Relación de la persona evaluada con el jefe/a de hogar"
            )
        with c4:
            programa_unidad = st.selectbox(
                "Programa / Unidad Responsable:",
                PROGRAMA_OPTIONS,
                key="programa_unidad",
                help="Programa o unidad del CESFAM que aplica la encuesta"
            )

        # Tipo de Unión (Metadata para Genograma)
        tipo_union = st.selectbox(
            "Tipo de Unión (Pareja Principal):",
            TIPO_UNION_OPTIONS,
            key="tipo_union",
            help="Define cómo se dibujará la línea de unión en el genograma (Guía Clínica)"
        )
        st.markdown('<hr style="border-top: 1px solid #e2e8f0; margin: 12px 0;">', unsafe_allow_html=True)
        st.markdown('<div style="font-weight:600; color: #334155; margin-bottom:4px;">📁 Carpeta Digital (Google Drive) — Referencia del Registro Físico:</div>', unsafe_allow_html=True)
        st.text_input(
            "Enlace Carpeta Drive",
            label_visibility="collapsed",
            placeholder="https://drive.google.com/drive/folders/...",
            key="link_drive",
            help="Ingrese el enlace a la carpeta Drive donde se almacena la pauta física escaneada de esta familia."
        )

    st.markdown("<br>", unsafe_allow_html=True)
    
    # --- 2. GRUPO FAMILIAR ---
    st.markdown('<div style="color: #0f172a; font-size: 1.1rem; font-weight: 700; margin-bottom: 8px;">2. Composición Familiar</div>', unsafe_allow_html=True)
    
    with st.container(border=True):
        # Ensure 'Pueblo Originario' column always exists for older records
        if 'Pueblo Originario' not in st.session_state.family_members.columns:
            if 'Identidad de género' in st.session_state.family_members.columns:
                loc = st.session_state.family_members.columns.get_loc('Identidad de género') + 1
            else:
                loc = min(3, len(st.session_state.family_members.columns))
            st.session_state.family_members.insert(loc, 'Pueblo Originario', "")
            
        # Ensure 'Parentesco' column exists
        if 'Parentesco' not in st.session_state.family_members.columns:
            loc = max(0, len(st.session_state.family_members.columns) - 1)
            st.session_state.family_members.insert(loc, 'Parentesco', "")

        if 'F. Nac' in st.session_state.family_members.columns:
            def to_date_safe_fnac(x):
                try:
                    return pd.to_datetime(x, dayfirst=True).date() if pd.notnull(x) and str(x).strip() != "" else None
                except:
                    return None
            st.session_state.family_members['F. Nac'] = st.session_state.family_members['F. Nac'].apply(to_date_safe_fnac)
        
        edited_family = st.data_editor(
            st.session_state.family_members,
            num_rows="dynamic",
            width='stretch',
            column_config={
                "Nombre y Apellidos": st.column_config.TextColumn("Nombre y Apellidos", width="large"),
                "RUT": st.column_config.TextColumn("RUT", width="medium"),
                "F. Nac": st.column_config.DateColumn("F. Nac", width="small", format="DD/MM/YYYY"),
                "Identidad de género": st.column_config.SelectboxColumn(
                    "Identidad de género", 
                    options=["Masculino", "Femenino", "No binario", "Transgénero", "Prefiero no decir", "Gestación/Aborto"], 
                    width="medium", 
                    help="Identidad inclusiva. Use 'Gestación/Aborto' para gestaciones (Triángulo)."
                ),
                "Pueblo Originario": st.column_config.SelectboxColumn(
                    "Etnia",
                    options=PUEBLO_ORIGINARIO_OPTIONS,
                    width="medium",
                    help="Autoidentificación étnica según INE Chile (Censo 2017)."
                ),
                "Nacionalidad": st.column_config.TextColumn("Nacionalidad", width="medium"),
                "E. Civil": st.column_config.SelectboxColumn(
                    "E. Civil", 
                    options=["Soltero/a (S)", "Casado/a (C)", "Conviviente (Co)", "Divorciado/a (D)", "Separado/a (Sep)", "Viudo/a (V)", "Fallecido/a (F)", "Espontáneo", "Provocado"], 
                    width="medium",
                    help="Use abreviaturas tradicionales o elija el tipo de aborto para gestaciones."
                ),
                "Ocupación": st.column_config.TextColumn("Ocupación", width="medium"),
                "Parentesco": st.column_config.SelectboxColumn(
                    "Parentesco",
                    options=PARENTESCO_FAMILIA_OPTIONS,
                    width="medium",
                    help="Especifique si es Gemelo (Fraterno/Idéntico) para dibujo especial."
                ),
                "Cronico": st.column_config.CheckboxColumn("Cron.", width="small", default=False, help="Marque si tiene Enfermedad Crónica (Borde Rojo)."),
                "Resp": st.column_config.CheckboxColumn("Resp", width="small", default=False, help="Marque si es la Persona Índice (Doble Borde)."),
            }
        )
        st.session_state.family_members = edited_family
        
        # --- VALIDACIÓN DE RUTS DUPLICADOS ---
        all_ruts = get_all_ruts_mapping()
        current_eval_id = st.session_state.get('idEvaluacion', '')
        
        dupes_found = []
        for idx, m_row in edited_family.iterrows():
            m_rut = str(m_row.get("RUT", "")).strip().upper()
            if m_rut in all_ruts:
                fam_name, other_id = all_ruts[m_rut]
                if other_id != current_eval_id:
                    dupes_found.append(f"• **{m_rut}** ({m_row.get('Nombre y Apellidos', 'Sin nombre')}) ya existe en la familia: **{fam_name}** (ID: {other_id})")
        
        if dupes_found:
            st.warning("⚠️ **Alerta de Duplicidad detectada:**\n\n" + "\n".join(dupes_found))

        st.markdown("""
<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:12px 16px;font-size:0.8rem;color:#1e3a5f;line-height:1.7">
<b>Composición Familiar — Guía de Campos</b><br>
<b>GÉNERO:</b> Masculino | Femenino | No binario | Transgénero | Prefiero no decir | Gestación/Aborto<br>
<b>ETNIA (INE):</b> Ninguno | Mapuche | Aymara | Rapa Nui | Atacameño (Lickanantay) | Quechua | Colla | Diaguita | Kawésqar | Yagán | Changos | Afrodescendiente | Otro<br>
<b>PARENTESCO:</b> Jefe/a de Hogar | Cónyuge/Pareja | Hijo/a | Gemelo Fraterno/Idéntico | Padre/Madre | Hermano/a | Abuelo/a | Nieto/a | Tío/a | Sobrino/a | Adoptivo/a | Otro familiar | No familiar<br>
<b>E. CIVIL:</b> S=Soltero/a &nbsp; C=Casado/a &nbsp; Co=Conviviente &nbsp; D=Divorciado/a &nbsp; Sep=Separado/a &nbsp; V=Viudo/a &nbsp; F=Fallecido/a<br>
<b>GESTACIÓN:</b> Identidad='Gestación/Aborto' + E.Civil vacío (en curso) | E.Civil='Espontáneo' (▵ con X) | E.Civil='Provocado' (▵ relleno)<br>
<b>GÉNEROS INCLUSIVOS</b> (No binario / Transgénero / Prefiero no decir): dibujan <b>Rombo (◇)</b> en el Genograma.
</div>""", unsafe_allow_html=True)


    st.markdown("<br>", unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # --- 3. PAUTA DE EVALUACIÓN ---
    st.markdown('<div style="color: #0f172a; font-size: 1.1rem; font-weight: 700; margin-bottom: 8px;">3. Pauta de Riesgo (Bíopsicosocial)</div>', unsafe_allow_html=True)
    
    with st.container(border=True):
        st.caption("Marque todos los factores de riesgo o protectores presentes en la familia evaluada.")
        
        # TABLA 1
        st.markdown('<div style="background: #fee2e2; color: #7f1d1d; padding: 8px 12px; border-radius: 8px; font-weight: 700; font-size: 0.85rem; margin: 16px 0 8px 0;">🔴 RIESGO MÁXIMO (No otorgan pts, clasifican directo a ALTO)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.checkbox("Familia con VIF (física, psicológica, sexual, económica)", key='t1_vif')
            st.checkbox("Consumo problema de drogas o dependencia", key='t1_drogas')
            st.checkbox("Consumo problema de alcohol (AUDIT > 13)", key='t1_alcohol')
            st.checkbox("Patología salud mental descompensada o sin TTO", key='t1_saludMentalDescomp')
            st.checkbox("Abuso sexual (sufrido por algún integrante)", key='t1_abusoSexual')
        with c2:
            st.checkbox("Adulto mayor y/o niño/a en riesgo biopsicosocial grave", key='t1_riesgoBiopsicoGrave')
            st.checkbox("Pauta EPSA (ChCC) con riesgo", key='t1_epsaRiesgo')
            st.checkbox("Vulnerabilidad socioeconómica extrema (indigencia)", key='t1_vulnerabilidadExtrema')
            st.checkbox("Trabajo infantil en niños < 14 años", key='t1_trabajoInfantil')

        # TABLA 2
        st.markdown('<div style="background: #ffedd5; color: #9a3412; padding: 8px 12px; border-radius: 8px; font-weight: 700; font-size: 0.85rem; margin: 24px 0 8px 0;">🟠 RIESGO ALTO (No otorgan pts, 2 factores = ALTO)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.checkbox("Enfermedad grave o terminal integrante", key='t2_enfermedadGrave')
            st.checkbox("Paciente con alto riesgo de hospitalizar", key='t2_altoRiesgoHosp')
            st.checkbox("Discapacidad física y/o mental (Bartel 35 o menos)", key='t2_discapacidad')
            st.checkbox("Patología de salud mental leve o moderada", key='t2_saludMentalLeve')
        with c2:
            st.checkbox("Conflictos o problemas con la justicia", key='t2_judicial')
            st.checkbox("Incumplimiento de roles parentales", key='t2_rolesParentales')
            st.checkbox("Sobrecarga del cuidador principal", key='t2_sobrecargaCuidador')
            st.checkbox("Conflictos familiares severos o crisis de comunicación", key='t2_conflictosSeveros')
            st.checkbox("Adultos en riesgo biopsicosocial a cargo de niños", key='t2_adultosRiesgo')

        # TABLA 3
        st.markdown('<div style="background: #fef9c3; color: #854d0e; padding: 8px 12px; border-radius: 8px; font-weight: 700; font-size: 0.85rem; margin: 24px 0 8px 0;">🟡 RIESGO MEDIO (4 pts. c/u)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.checkbox("Patología crónica descompensada sintomática", key='t3_patologiaCronica')
            st.checkbox("Miembro con discapacidad leve/moderada (40-55pts)", key='t3_discapacidadLeve')
            st.checkbox("Rezago desarrollo psicomotor", key='t3_rezago')
            st.checkbox("Madre adolescente", key='t3_madreAdolescente')
            st.checkbox("Duelo reciente (pérdida de integrante significativo)", key='t3_duelo')
            st.checkbox("Ausencia o escasa red de apoyo social/familiar", key='t3_sinRedApoyo')
            st.checkbox("Cesantía de más de 1 mes del proveedor", key='t3_cesantia')
            st.checkbox("Vulnerabilidad socioeconómica no extrema", key='t3_vulneNoExtrema')
            st.checkbox("Precariedad laboral (temporal/honorarios)", key='t3_precariedadLaboral')
        with c2:
            st.checkbox("Hacinamiento (2.5+ personas por dormitorio)", key='t3_hacinamiento')
            st.checkbox("Entorno inseguro (delincuencia)", key='t3_entornoInseguro')
            st.checkbox("Adulto mayor que vive solo", key='t3_adultoSolo')
            st.checkbox("Deserción o fracaso escolar", key='t3_desercionEscolar')
            st.checkbox("Analfabetismo padre/madre/cuidador", key='t3_analfabetismo')
            st.checkbox("Escolaridad básica incompleta padres", key='t3_escolaridadIncompleta')
            st.checkbox("Dificultad de acceso a servicios", key='t3_dificultadAcceso')

        # TABLA 4
        st.markdown('<div style="background: #e0f2fe; color: #075985; padding: 8px 12px; border-radius: 8px; font-weight: 700; font-size: 0.85rem; margin: 24px 0 8px 0;">🔵 RIESGO BAJO (3 pts. c/u)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.checkbox("Hogar monoparental", key='t4_monoparental')
            st.checkbox("Riesgo cardiovascular (tabaco, obesidad)", key='t4_riesgoCardio')
            st.checkbox("Foco contaminación ambiental cercano", key='t4_contaminacion')
            st.checkbox("Deficiencia hábitos higiene", key='t4_higiene')
        with c2:
            st.checkbox("Ausencia prácticas recreación", key='t4_sinRecreacion')
            st.checkbox("Ausencia espacios seguros recreación", key='t4_sinEspaciosSeguros')
            st.checkbox("Endeudamiento familiar elevado (>40%)", key='t4_endeudamiento')
            st.checkbox("Servicios básicos incompletos/inadecuados", key='t4_serviciosIncompletos')

        # TABLA 5
        st.markdown('<div style="background: #dcfce7; color: #166534; padding: 8px 12px; border-radius: 8px; font-weight: 700; font-size: 0.85rem; margin: 24px 0 8px 0;">🟢 FACTORES PROTECTORES (No otorgan pts)</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="medium")
        with c1:
            st.checkbox("Lactancia materna exclusiva/complementaria", key='t5_lactancia')
            st.checkbox("Hábitos saludables (actividad física, alim.)", key='t5_habitos')
            st.checkbox("Presencia redes sociales/comunitarias", key='t5_redesSociales')
            st.checkbox("Presencia red apoyo familiar", key='t5_redFamiliar')
        with c2:
            st.checkbox("Habilidades comunicacionales (afecto)", key='t5_comunicacion')
            st.checkbox("Recursos socioeconómicos suficientes", key='t5_recursosSuficientes')
            st.checkbox("Resiliencia (sobreponerse a crisis)", key='t5_resiliencia')
            st.checkbox("Vivienda adecuada", key='t5_viviendaAdecuada')

        # --- SECCIÓN APGAR FAMILIAR ---
        st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 24px 0;'>", unsafe_allow_html=True)
        st.markdown('<div style="background: #f1f5f9; padding: 12px; border-radius: 8px; font-weight: 700; font-size: 0.9rem; margin-bottom: 12px;">📊 APGAR FAMILIAR (Funcionamiento Familiar)</div>', unsafe_allow_html=True)
        
        apgar_options = {
            0: "0 - Casi nunca",
            1: "1 - A veces",
            2: "2 - Casi siempre"
        }
        
        c_ap1, c_ap2 = st.columns(2, gap="medium")
        with c_ap1:
            a1 = st.selectbox("Adaptación: ¿Está satisfecho con la ayuda que recibe de su familia cuando tiene problemas?", options=[0,1,2], format_func=lambda x: apgar_options[x], key="apgar_a1")
            a2 = st.selectbox("Participación: ¿Está satisfecho con la forma en que su familia discute problemas y comparte soluciones?", options=[0,1,2], format_func=lambda x: apgar_options[x], key="apgar_a2")
            a3 = st.selectbox("Graduación/Crecimiento: ¿Siente que su familia acepta y apoya sus nuevos intereses o cambios?", options=[0,1,2], format_func=lambda x: apgar_options[x], key="apgar_a3")
        with c_ap2:
            a4 = st.selectbox("Afecto: ¿Está satisfecho con la forma en que su familia expresa afecto y responde a sus emociones?", options=[0,1,2], format_func=lambda x: apgar_options[x], key="apgar_a4")
            a5 = st.selectbox("Resolución: ¿Está satisfecho con la cantidad de tiempo que comparte con su familia?", options=[0,1,2], format_func=lambda x: apgar_options[x], key="apgar_a5")
            
        apgar_total = a1 + a2 + a3 + a4 + a5
        st.session_state['apgar_total'] = apgar_total
        
        if apgar_total >= 7:
            apgar_label = "FAMILIA FUNCIONAL (7-10 pts)"
            apgar_color = "#166534"
        elif 4 <= apgar_total <= 6:
            apgar_label = "DISFUNCIÓN LEVE (4-6 pts)"
            apgar_color = "#854d0e"
        else:
            apgar_label = "DISFUNCIÓN SEVERA (0-3 pts)"
            apgar_color = "#991b1b"
            
        st.markdown(f'<div style="text-align: right; font-weight: 700; color: {apgar_color};">Resultado APGAR: {apgar_label}</div>', unsafe_allow_html=True)

    # CÁLCULO
    count_t1 = sum([st.session_state[k] for k in risk_keys if k.startswith('t1_')])
    count_t2 = sum([st.session_state[k] for k in risk_keys if k.startswith('t2_')])
    count_t3 = sum([st.session_state[k] for k in risk_keys if k.startswith('t3_')])
    count_t4 = sum([st.session_state[k] for k in risk_keys if k.startswith('t4_')])

    score_medium = count_t3 * 4
    score_low = count_t4 * 3
    total_points = score_medium + score_low

    level = 'SIN RIESGO'
    border_color = 'green'
    text_color = 'green'
    
    if count_t1 >= 1 or count_t2 >= 2 or total_points >= 26:
        level = 'RIESGO ALTO'
        border_color = '#d32f2f'
        text_color = '#d32f2f'
    elif (total_points >= 17 and total_points <= 25) or count_t2 == 1:
        level = 'RIESGO MEDIO'
        border_color = '#ed6c02'
        text_color = '#ed6c02'
    else:
        level = 'RIESGO BAJO'
        border_color = '#2e7d32'
        text_color = '#2e7d32'

    # --- SAVE CALCULATIONS TO SESSION STATE FOR PDF ---
    st.session_state['count_t1'] = count_t1
    st.session_state['count_t2'] = count_t2
    st.session_state['count_t3'] = count_t3
    st.session_state['count_t4'] = count_t4
    st.session_state['score_medium'] = score_medium
    st.session_state['score_low'] = score_low
    st.session_state['total_points'] = total_points
    st.session_state['level'] = level

    # RESULTADOS
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background: white; border-radius: 16px; padding: 24px; box-shadow: 0 4px 20px rgba(0,0,0,0.05); border: 1px solid #f3f4f6; display:flex; justify-content:space-between; align-items:center; font-family: 'Inter', sans-serif;">
        <div style="font-size: 0.85rem; color: #64748b; line-height: 1.6;">
            <p style="margin:0;"><strong style="color:#ef4444;">RIESGO ALTO:</strong> 1 factor máx, o 2 altos, o desde 26pts.</p>
            <p style="margin:0;"><strong style="color:#f59e0b;">RIESGO MEDIO:</strong> 17-25 pts, o solo 1 factor alto.</p>
            <p style="margin:0;"><strong style="color:#10b981;">RIESGO BAJO:</strong> Hasta 16 pts.</p>
        </div>
        <div style="text-align:right;">
            <div style="border-bottom: 1px solid #f1f5f9; margin-bottom:8px; padding-bottom:8px;">
                <span style="font-weight:600; font-size: 0.9rem; color: #475569;">PUNTAJE OBTENIDO:</span>
                <span style="font-size: 1.8rem; font-weight:800; color: #0f172a; margin-left:12px;">{total_points}</span>
            </div>
            <div>
                <span style="font-weight:600; font-size: 0.9rem; color: #475569;">NIVEL DE RIESGO:</span>
                <span style="font-size: 1.5rem; font-weight:800; color:{text_color}; margin-left:12px; text-transform:uppercase;">{level}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Firma Evaluador
    st.markdown("<br>", unsafe_allow_html=True)
    c_eval_1, c_eval_2, c_eval_3 = st.columns([1, 2, 1]) 
    with c_eval_2:
        st.text_input("Firma Evaluador Input", label_visibility="collapsed", key="sig_evaluador_input")
        st.markdown('<div style="text-align:center; font-weight:600; font-size:0.8rem; color:#64748b; padding-top: 5px;">NOMBRE Y FIRMA DEL EVALUADOR</div>', unsafe_allow_html=True)

    # --- 4. PLAN INTERVENCIÓN y FIRMAS ---
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div style="color: #0f172a; font-size: 1.1rem; font-weight: 700; margin-bottom: 8px;">4. Plan de Intervención y Compromisos</div>', unsafe_allow_html=True)
    
    with st.container(border=True):
        edited_plan = st.data_editor(
            st.session_state.intervention_plan,
            num_rows="dynamic",
            width='stretch',
            column_config={
                "Objetivo": st.column_config.TextColumn("OBJETIVO"),
                "Actividad": st.column_config.TextColumn("ACTIVIDAD"),
                "Fecha Prog": st.column_config.DateColumn("FECHA PROG."),
                "Responsable": st.column_config.TextColumn("RESPONSABLE"),
                "Fecha Real": st.column_config.DateColumn("FECHA REAL."),
                "Evaluación": st.column_config.TextColumn("EVALUACIÓN"),
                "Estado": st.column_config.SelectboxColumn(
                    "ESTADO",
                    options=["Pendiente", "En progreso", "Completado", "Cancelado"],
                    width="medium",
                    help="Estado actual de la actividad."
                ),
                "F. Seguimiento": st.column_config.DateColumn("F. SEGUIMIENTO"),
                "Obs. Seguimiento": st.column_config.TextColumn("OBS. SEGUIMIENTO"),
            }
        )

        if edited_plan is not None:
            for col in ["Fecha Prog", "Fecha Real"]:
                if col in edited_plan.columns:
                     def to_date_safe(x):
                         try:
                             return pd.to_datetime(x).date() if pd.notnull(x) else None
                         except:
                             return None
                     edited_plan[col] = edited_plan[col].apply(to_date_safe)
            
            st.session_state.intervention_plan = edited_plan

        st.markdown('<div style="color: #64748b; font-size: 0.8rem; font-style: italic; margin-top: 8px;">* Si se han considerado otros elementos diagnósticos en la evaluación, favor detallar en observaciones.</div>', unsafe_allow_html=True)
        
        st.markdown('<div style="margin-top:16px; font-weight:600; color: #334155;">Observaciones Clínicas:</div>', unsafe_allow_html=True)
        st.text_area("Observaciones", label_visibility="collapsed", height=80, key="observaciones")
        

        st.markdown('<hr style="border-top: 1px solid #e2e8f0; margin: 20px 0;">', unsafe_allow_html=True)
        st.markdown('<div style="color: #0f172a; font-size: 1rem; font-weight: 700; margin-bottom: 8px;">📊 Seguimiento del Plan de Intervención</div>', unsafe_allow_html=True)
        st.caption("Registre el avance de cada actividad del plan. Esta tabla se guarda junto la evaluación para auditoría.")

        if 'seguimiento_plan' not in st.session_state:
            st.session_state.seguimiento_plan = pd.DataFrame({
                "Objetivo": pd.Series(dtype='str'),
                "Actividad": pd.Series(dtype='str'),
                "Estado": pd.Series(dtype='str'),
                "F. Seguimiento": pd.Series(dtype='object'),
                "Obs. Seguimiento": pd.Series(dtype='str'),
            })
        # Ensure columns always exist
        for _sc in ['Objetivo', 'Actividad', 'Estado', 'F. Seguimiento', 'Obs. Seguimiento']:
            if _sc not in st.session_state.seguimiento_plan.columns:
                st.session_state.seguimiento_plan[_sc] = ''

        edited_seg = st.data_editor(
            st.session_state.seguimiento_plan,
            num_rows="dynamic",
            width='stretch',
            key="seguimiento_editor",
            column_config={
                "Objetivo": st.column_config.TextColumn("OBJETIVO", width="large"),
                "Actividad": st.column_config.TextColumn("ACTIVIDAD", width="large"),
                "Estado": st.column_config.SelectboxColumn(
                    "ESTADO",
                    options=["Pendiente", "En progreso", "Completado", "Cancelado"],
                    width="medium",
                ),
                "F. Seguimiento": st.column_config.DateColumn("FECHA SEGUIMIENTO", width="medium"),
                "Obs. Seguimiento": st.column_config.TextColumn("OBSERVACIONES", width="large"),
            }
        )
        st.session_state.seguimiento_plan = edited_seg
        
        # Firmas Equipo Salud 
        st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 24px 0;'>", unsafe_allow_html=True)
        c_f1, c_f2 = st.columns(2, gap="large")
        with c_f1:
            st.markdown('<div style="font-weight:600; font-size:0.9rem; color: #0f172a; margin-bottom:8px;">Equipo Salud Participante:</div>', unsafe_allow_html=True)
            edited_team = st.data_editor(
                st.session_state.team_members,
                num_rows="dynamic",
                width='stretch',
                column_config={
                    "Nombre y Profesión": st.column_config.TextColumn("Registrar Nombre Completo", width="large"),
                    "Firma": st.column_config.CheckboxColumn("Firma Digital", width="small"),
                }
            )
            st.session_state.team_members = edited_team
        with c_f2:
            st.markdown('<div style="font-weight:600; font-size:0.9rem; color: #0f172a; margin-bottom:8px;">Jefe Equipo de Cabecera:</div>', unsafe_allow_html=True)
            st.text_input("Nombre Jefe", label_visibility="collapsed", key="sig_jefe")
            st.markdown('<div style="font-size:0.8rem; color:#64748b;">(Firma Responsable Sector)</div>', unsafe_allow_html=True)
        
        # Compromiso Section
        st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 24px 0;'>", unsafe_allow_html=True)
        st.markdown('<div style="background: #f8fafc; padding: 16px; border-radius: 12px; border: 1px solid #e2e8f0;">', unsafe_allow_html=True)
        st.markdown('<div style="font-weight:700; color: #0f172a; margin-bottom: 12px; font-size: 0.95rem;">COMPROMISO CONJUNTO SALUD - FAMILIA</div>', unsafe_allow_html=True)
        
        c_L1_1, c_L1_2, c_L1_3, c_L1_4, c_L1_5 = st.columns([2.5, 2, 2.5, 3, 1.5])
        with c_L1_1:
             st.markdown('<div style="margin-top: 8px; font-size: 0.9rem;">El equipo de cabecera del sector</div>', unsafe_allow_html=True)
        with c_L1_2:
             sector_comp = st.selectbox("Sector Comp", options=["Sol", "Luna", "No identificado"], label_visibility="collapsed", key="comp_sector")
        with c_L1_3:
             st.markdown('<div style="margin-top: 8px; font-size: 0.9rem; text-align:right;">representado por don(ña)</div>', unsafe_allow_html=True)
        with c_L1_4:
             rep_sector = st.text_input("Rep Sector", label_visibility="collapsed", key="comp_rep_sector")
        with c_L1_5:
             st.markdown('<div style="margin-top: 8px; font-size: 0.9rem;">(Cargo),</div>', unsafe_allow_html=True)
        
        c_L2_1, c_L2_2, c_L2_3, c_L2_4 = st.columns([1, 3, 1.5, 4])
        with c_L2_1:
             st.markdown('<div style="margin-top: 8px; font-size: 0.9rem;">y la familia</div>', unsafe_allow_html=True)
        with c_L2_2:
             familia_comp = st.text_input("Familia Comp", label_visibility="collapsed", key="comp_familia")
        with c_L2_3:
             st.markdown('<div style="margin-top: 8px; font-size: 0.9rem; text-align:right;">domiciliada en</div>', unsafe_allow_html=True)
        with c_L2_4:
             dir_comp = st.text_input("Dir Comp", label_visibility="collapsed", key="comp_dir")

        c_L3_1, c_L3_2, c_L3_3, c_L3_4 = st.columns([2, 3, 2, 2])
        with c_L3_1:
            st.markdown('<div style="margin-top: 8px; font-size: 0.9rem;">representado por don(ña)</div>', unsafe_allow_html=True)
        with c_L3_2:
            rep_fam = st.text_input("Rep Familia", label_visibility="collapsed", key="comp_rep_fam")
        with c_L3_3:
            st.markdown('<div style="margin-top: 8px; font-size: 0.9rem; text-align:right;">RUT N°</div>', unsafe_allow_html=True)
        with c_L3_4:
            rut_rep = st.text_input("RUT Rep", label_visibility="collapsed", key="comp_rut")
        
        st.markdown('<div style="margin-top:16px; margin-bottom:8px; font-size: 0.9rem; line-height: 1.5;">Mediante el presente documento manifiestan el acuerdo de ejecutar el Plan de Trabajo elaborado en conjunto con el equipo de salud, con fecha de entrada en vigencia:</div>', unsafe_allow_html=True)
        fecha_comp = st.date_input("Fecha Comp", value=date.today(), key="comp_fecha")

        st.markdown("<br><br>", unsafe_allow_html=True)
        c_cf1, c_cf2 = st.columns(2, gap="large")
        with c_cf1:
            firma_func = st.text_input("Nombre Funcionario Firmante", label_visibility="collapsed", key="sig_func")
            st.markdown('<div style="text-align:center; font-weight:600; color:#64748b; font-size:0.8rem; padding-top: 4px;">Nombre Funcionario Firmante</div>', unsafe_allow_html=True)
        with c_cf2:
            firma_benef = st.text_input("Beneficiario Firmante", label_visibility="collapsed", key="sig_benef")
            st.markdown('<div style="text-align:center; font-weight:600; color:#64748b; font-size:0.8rem; padding-top: 4px;">Beneficiario Firmante</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True) # Cierra contenedor gris claro de compromiso
        
        st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 24px 0;'>", unsafe_allow_html=True)
        col_ft1, col_ft2 = st.columns([1, 2])
        with col_ft1:
            fecha_egreso = st.date_input("FECHA EGRESO PLAN:", value=None, key="fechaEgreso")
        with col_ft2:
            st.markdown("**Estado Egreso Validado:**")
            cf1, cf2, cf3, cf4 = st.columns(4)
            alta = cf1.checkbox("Alta", key="egreso_alta")
            traslado = cf2.checkbox("Traslado fam.", key="egreso_traslado")
            derivacion = cf3.checkbox("Derivación", key="egreso_derivacion")
            abandono = cf4.checkbox("Abandono", key="egreso_abandono")
    
    st.markdown("<br><br>", unsafe_allow_html=True)
    evaluador_nombre = st.text_input("Nombre Evaluador (para registro digital):", key="evaluadorName")
    
    # ---- BOTONES GUARDAR Y DESCARGAR ----
    col_save, col_down = st.columns([3, 1])
    
    with col_save:
        # Determinar si es registro nuevo o existente
        id_evaluacion = st.session_state.get('idEvaluacion', '')
        _es_registro_existente = bool(id_evaluacion)
        btn_label = "🔄 ACTUALIZAR REGISTRO" if _es_registro_existente else "💾 GUARDAR REGISTRO DIGITAL"
        if st.button(btn_label, width='stretch', type="primary"):
            with st.spinner("Guardando en la nube..."):
                # Si es registro nuevo, generar el ID ahora con el apellido ingresado
                if not id_evaluacion:
                    familia_para_id = st.session_state.get('familia', '')
                    id_evaluacion = generate_incremental_eval_id(familia_para_id)
                    st.session_state['idEvaluacion'] = id_evaluacion

                # Leer valores actuales desde session_state
                _parentesco = st.session_state.get('parentesco', '')
                _programa = st.session_state.get('programa_unidad', '')

                # ---- EXTRAER RUTs DEL GRUPO FAMILIAR (sin puntos: 1234567-8) ----
                def normalizar_rut(rut_str):
                    """Elimina puntos del RUT, conserva solo guión como separador."""
                    return rut_str.replace(".", "").strip()

                df_fam_rut = st.session_state.family_members.fillna("")
                if 'RUT' in df_fam_rut.columns:
                    ruts_list = [normalizar_rut(str(r)) for r in df_fam_rut['RUT'].tolist() if str(r).strip()]
                    ruts_concatenados = ",".join(ruts_list)
                else:
                    ruts_concatenados = ""

                # ---- HOJA 1: EVALUACIONES ----
                data_row = [
                    id_evaluacion,

                    str(fecha_input),
                    familia,
                    direccion,
                    establecimiento,
                    sector,
                    _parentesco,
                    _programa,
                    total_points,
                    level,
                    evaluador_nombre,
                    st.session_state.get('tipo_union', 'Casados'), # Tipo Unión
                    ruts_concatenados,   # RUTs del grupo familiar
                ]
                
                risk_list = [st.session_state.get(k, False) for k in risk_keys]
                data_row.extend(risk_list)

                df_fam = st.session_state.family_members.fillna("")
                family_json = json.dumps(df_fam.to_dict('records'), ensure_ascii=False, default=str)
                
                df_plan_save = st.session_state.intervention_plan.copy()
                for c in ['Fecha Prog', 'Fecha Real']:
                    if c in df_plan_save.columns:
                        df_plan_save[c] = df_plan_save[c].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notnull(x) and hasattr(x, 'strftime') else "")
                plan_json = json.dumps(df_plan_save.to_dict('records'), ensure_ascii=False)
                
                df_team = st.session_state.team_members.fillna("")
                team_json = json.dumps(df_team.to_dict('records'), ensure_ascii=False, default=str)
                
                rel_json = json.dumps(st.session_state.get('interpersonal_relations', []), ensure_ascii=False)
                
                # Seguimiento del Plan
                df_seg_save = st.session_state.get('seguimiento_plan', pd.DataFrame()).copy()
                seg_json = json.dumps(df_seg_save.fillna('').to_dict('records'), ensure_ascii=False, default=str)
                
                # Extra data (APGAR)
                apgar_val = st.session_state.get('apgar_total', 0)
                a1 = st.session_state.get('apgar_a1', 0)
                a2 = st.session_state.get('apgar_a2', 0)
                a3 = st.session_state.get('apgar_a3', 0)
                a4 = st.session_state.get('apgar_a4', 0)
                a5 = st.session_state.get('apgar_a5', 0)
                
                data_row.append(family_json)
                data_row.append(plan_json)
                data_row.append(team_json)
                data_row.append(rel_json)
                data_row.append(seg_json)
                data_row.append(apgar_val)
                data_row.append(a1)
                data_row.append(a2)
                data_row.append(a3)
                data_row.append(a4)
                data_row.append(a5)
                
                extra_data = [
                    st.session_state.get('comp_rep_sector', ''),
                    st.session_state.get('comp_familia', ''),
                    st.session_state.get('comp_dir', ''),
                    st.session_state.get('comp_rep_fam', ''),
                    st.session_state.get('comp_rut', ''),
                    str(st.session_state.get('comp_fecha', date.today())),
                    st.session_state.get('sig_func', ''),
                    st.session_state.get('sig_benef', ''),
                    st.session_state.get('sig_equipo', ''),
                    st.session_state.get('sig_jefe', ''),
                    st.session_state.get('sig_evaluador_input', ''),
                    st.session_state.get('egreso_alta', False),
                    st.session_state.get('egreso_traslado', False),
                    st.session_state.get('egreso_derivacion', False),
                    st.session_state.get('egreso_abandono', False),
                    str(st.session_state.get('fechaEgreso', '')) if st.session_state.get('fechaEgreso') else "",
                    st.session_state.get('observaciones', ''),
                    st.session_state.get('link_drive', '')
                ]
                data_row.extend(extra_data)

                final_headers = [
                    "ID Evaluación", "Fecha", "Familia", "Dirección", "Establecimiento", "Sector",
                    "Parentesco", "Programa/Unidad",
                    "Puntaje", "Nivel", "Evaluador",
                    "Tipo Unión", "RUTs Grupo Familiar",
                ] + risk_keys + [
                    "Grupo Familiar JSON", "Plan Intervención JSON", "Equipo Salud JSON", "Relaciones JSON", "Seguimiento Plan JSON",
                    "APGAR Total", "A1", "A2", "A3", "A4", "A5"
                ] + [
                    "Rep Sector", "Familia Comp", "Dir Comp", "Rep Familia", "RUT Rep", "Fecha Comp",
                    "Firma Funcionario", "Firma Beneficiario", "Firma Equipo", "Firma Jefe", "Firma Evaluador",
                    "egreso_alta", "egreso_traslado", "egreso_derivacion", "egreso_abandono",
                    "Fecha Egreso", "Observaciones", "Carpeta Digital (Drive)"
                ]
                
                success1, msg1 = save_evaluacion_to_sheet(data_row, final_headers)
                
                # ---- HOJA 2: PLANES DE INTERVENCIÓN ----
                success2, msg2 = save_intervention_rows(
                    id_evaluacion,
                    familia,
                    str(fecha_input),
                    level,
                    _programa,
                    _parentesco,
                    st.session_state.intervention_plan
                )

                # ---- HOJA 3: REM-P7 (auto-actualizar) ----
                _sol_ins  = st.session_state.get('n_inscritas_sol', 0)
                _luna_ins = st.session_state.get('n_inscritas_luna', 0)
                success3, msg3 = update_rem_p7(_sol_ins, _luna_ins)
                
                if success1 and success2:
                    rem_info = f"\n\n📊 REM-P7: {msg3}" if success3 else f"\n\n⚠️ REM-P7: {msg3}"
                    st.success(f"✅ Ficha guardada exitosamente.\n\n📋 Evaluación: {msg1}\n\n📌 Plan: {msg2}{rem_info}")
                    st.balloons()
                elif success1:
                    st.warning(f"⚠️ Evaluación guardada, pero hubo un problema con el Plan: {msg2}")
                else:
                    st.error(f"❌ Error al guardar evaluación: {msg1}")

    with col_down:
        if st.button("📄 Preparar PDF Evaluación", width='stretch'):
            try:
                id_evaluacion = st.session_state.get('idEvaluacion', 'sin_id')
                with st.spinner("Preparando archivo PDF..."):
                    pdf_bytes = generate_pdf_report(
                        dict(st.session_state), 
                        st.session_state.family_members, 
                        st.session_state.intervention_plan,
                        st.session_state.team_members
                    ) 
                    st.session_state['temp_pdf_report'] = pdf_bytes
                    # Auditoría PDF
                    log_audit_event(st.session_state.user_info, "Generación de PDF", f"PDF preparado para la familia: {st.session_state.get('familia', 'N/A')}", eval_id=id_evaluacion)
                    st.success("✅ PDF listo")
            except Exception as e:
                st.error(f"Error generando PDF: {e}")

        if 'temp_pdf_report' in st.session_state:
            id_eval_filename = st.session_state.get('idEvaluacion', 'sin_id')
            st.download_button(
                label="⬇️ Descargar archivo PDF",
                data=st.session_state['temp_pdf_report'],
                file_name=f"ficha_oficial_{id_eval_filename}.pdf",
                mime="application/pdf",
                width='stretch',
            )

    # --- FOOTER PROFESIONAL (dentro de main - sólo usuarios autenticados ven la app) ---
    st.markdown("---")
    col_f1, col_f2, col_f3 = st.columns([1, 2, 1])
    with col_f2:
        st.markdown("""
            <div style='
                text-align: center;
                padding: 18px 24px;
                background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
                border-radius: 12px;
                border-top: 3px solid #1e3a8a;
                margin-top: 10px;
            '>
                <div style='font-size: 15px; font-weight: 700; color: #1e3a8a; letter-spacing: 0.5px;'>
                    🏥 Jefatura Técnica — CESFAM Cholchol
                </div>
                <div style='height: 2px; width: 40px; background: #fbbf24; margin: 8px auto;'></div>
                <div style='font-size: 12px; color: #64748b; margin-top: 6px; line-height: 1.8;'>
                    💼 Desarrollado por <strong>Alain Antinao Sepúlveda</strong><br>
                    📧 <a href="mailto:alain.antinao.s@gmail.com" style="color: #2563eb; text-decoration: none;">alain.antinao.s@gmail.com</a>
                    &nbsp;·&nbsp;
                    🌐 <a href="https://alain-antinao-s.notion.site/Alain-C-sar-Antinao-Sep-lveda-1d20a081d9a980ca9d43e283a278053e"
                           target="_blank" style="color: #2563eb; text-decoration: none;">Página personal</a>
                </div>
            </div>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
