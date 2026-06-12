import pandas as pd
import gspread
import io
import datetime
import streamlit as st
import json

PERCAPITA_SHEET_URL = st.secrets["PERCAPITA_SHEET_URL"]
WORKSHEET_NAME = "percapita"

def map_month_to_num(month_str):
    """Convierte el nombre del mes en español a número (1-12)"""
    if not isinstance(month_str, str):
        return 0
    m = month_str.strip().lower()
    months = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    return months.get(m, 0)

@st.cache_data(ttl=600, show_spinner=False)
def fetch_percapita_data():
    """Descarga los datos percapita usando las credenciales del app.py"""
    try:
        from app import get_google_sheet_client
        client = get_google_sheet_client()
        if not client:
            return None, "Error de conexión con Google Sheets."
        
        spreadsheet = client.open_by_url(PERCAPITA_SHEET_URL)
        worksheet = spreadsheet.worksheet(WORKSHEET_NAME)
        
        all_data = worksheet.get_all_values()
        if len(all_data) < 2:
            return pd.DataFrame(), "Hoja percapita vacía."
        
        df = pd.DataFrame(all_data[1:], columns=all_data[0])
        return df, None
    except gspread.exceptions.APIError as e:
        if e.response.status_code in [403, 404]:
            return None, f"Acceso denegado a la planilla Percapita. Verifique que se ha compartido el libro '{PERCAPITA_SHEET_URL}' con el correo de la Service Account."
        return None, f"Error API de Google: {e}"
    except gspread.exceptions.WorksheetNotFound:
        return None, f"No se encontró la hoja llamada '{WORKSHEET_NAME}'."
    except Exception as e:
        return None, str(e)

def process_percapita_data(df):
    """Agrupa por el máximo mes y año."""
    if df.empty:
        return df, "Periodo no encontrado"
        
    required_cols = ['RUT', 'ANIO_CORTE', 'MES_CORTE']
    for col in required_cols:
        if col not in df.columns:
            return pd.DataFrame(), f"Falta la columna requerida: {col}"

    df_clean = df.copy()
    
    # Limpieza de valores vacíos y conversión a numérico
    df_clean = df_clean[df_clean['ANIO_CORTE'].str.strip() != ""]
    df_clean['ANIO_CORTE_NUM'] = pd.to_numeric(df_clean['ANIO_CORTE'], errors='coerce').fillna(0).astype(int)
    df_clean['MES_CORTE_NUM'] = df_clean['MES_CORTE'].apply(map_month_to_num)
    
    if df_clean.empty:
        return df_clean, "No hay datos válidos para procesar."
        
    # Identificar el mayor año y mes
    max_year = df_clean['ANIO_CORTE_NUM'].max()
    max_month = df_clean[df_clean['ANIO_CORTE_NUM'] == max_year]['MES_CORTE_NUM'].max()
    
    df_filtered = df_clean[(df_clean['ANIO_CORTE_NUM'] == max_year) & (df_clean['MES_CORTE_NUM'] == max_month)]
    
    period_str = f"Mes: {max_month} | Año: {max_year}"
    if not df_filtered.empty:
        # Intentar obtener el nombre del mes para el mensaje
        max_month_name = df_filtered.iloc[0].get('MES_CORTE', '')
        period_str = f"{str(max_month_name).capitalize()} {max_year}"
    
    return df_filtered, period_str

def cross_reference_families(df_percapita_filtered):
    """
    Lee todas las evaluaciones locales, extrae los integrantes del 'Grupo Familiar JSON'
    y los cruza con el DataFrame de percapitados.
    """
    from app import get_google_sheet_client, SHEET_URL
    import json
    
    client = get_google_sheet_client()
    if not client:
        return pd.DataFrame()
        
    try:
        sh = client.open_by_url(SHEET_URL)
        ws = sh.worksheet("Evaluaciones")
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return pd.DataFrame()
            
        df_evals = pd.DataFrame(all_vals[1:], columns=all_vals[0])
    except Exception:
        return pd.DataFrame()
        
    # Extraer todos los miembros
    members_data = []
    
    for _, row in df_evals.iterrows():
        fam_json = row.get("Grupo Familiar JSON", "[]")
        eval_id = row.get("ID Evaluación", "Desconocido")
        familia = row.get("Familia", "Desconocida")
        sector = row.get("Sector", "No Identificado")
        
        try:
            gf = json.loads(fam_json)
            for m in gf:
                rut_miembro = str(m.get("RUT", "")).strip().upper()
                if not rut_miembro or rut_miembro == "S/R":
                    continue
                
                members_data.append({
                    "ID Evaluación": eval_id,
                    "Familia": familia,
                    "Sector": sector,
                    "Nombre Miembro": str(m.get("Nombre y Apellidos", "")).strip(),
                    "RUT Miembro": rut_miembro,
                    "Parentesco": str(m.get("Parentesco", "")),
                })
        except:
            continue
            
    df_members = pd.DataFrame(members_data)
    if df_members.empty:
        return df_members

    # Limpiar RUTs en Percapita para un cruce exacto
    def clean_rut(rut_str):
        if not isinstance(rut_str, str):
            return str(rut_str).upper()
        # Remueve puntos, permite guion, todo mayúscula
        import re
        r = rut_str.replace(".", "").strip().upper()
        # Asegurarse de que si el RUT tiene formato sin guion, intentar estandarizar
        # (Dependerá de cómo viene en la base percapita, asumimos que viene con guion o sin él pero sin puntos)
        return r

    df_percapita_filtered['RUT_CLEAN'] = df_percapita_filtered['RUT'].apply(clean_rut)
    df_members['RUT_CLEAN'] = df_members['RUT Miembro'].apply(clean_rut)
    
    # Cruce: Left join de Members con Percapita
    df_merged = pd.merge(
        df_members, 
        df_percapita_filtered[['RUT_CLEAN', 'NOMBRE_CENTRO']], 
        on='RUT_CLEAN', 
        how='left'
    )
    
    # Marcar percapitado si tiene cruce exitoso
    df_merged['Estado Percapita'] = df_merged['NOMBRE_CENTRO'].apply(lambda x: 'Percapitado' if pd.notnull(x) and x != "" else 'No Percapitado')
    df_merged['Centro Percapita'] = df_merged['NOMBRE_CENTRO'].fillna('No Registrado')
    
    # Reordenar columnas
    cols_to_show = ["ID Evaluación", "Familia", "Sector", "Nombre Miembro", "RUT Miembro", "Estado Percapita", "Centro Percapita"]
    return df_merged[cols_to_show]

def export_percapita_dashboard_excel(df_cruzado, periodo_str, user_info):
    """
    Genera el Excel del reporte de Percapita con hojas de Inicio, Dashboard y Base de Datos,
    incluyendo gráficos interactivos de análisis.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    import os
    import io
    from datetime import datetime
    from zoneinfo import ZoneInfo
    
    # Configurar zona horaria de Chile
    tz_chile = ZoneInfo('America/Santiago')
    fecha_generacion = datetime.now(tz_chile).strftime('%d/%m/%Y %H:%M')
    
    wb = Workbook()
    
    # Estilos globales
    DARK_BLUE = PatternFill("solid", fgColor="1F3864")
    CELESTE = PatternFill("solid", fgColor="BDD7EE")
    GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
    RED_FILL = PatternFill("solid", fgColor="FEE2E2")
    GRAY_FILL = PatternFill("solid", fgColor="F3F4F6")
    
    BOLD_WHITE = Font(bold=True, color="FFFFFF", size=12)
    BOLD_DARK = Font(bold=True, color="1F3864", size=11)
    BOLD_BLACK = Font(bold=True, color="000000", size=11)
    TITLE_FONT = Font(bold=True, color="1F3864", size=16)
    NORMAL = Font(size=10)
    GREEN_FONT = Font(color="166534", bold=True)
    RED_FONT = Font(color="991B1B", bold=True)
    
    THIN = Side(style="thin", color="B8CCE4")
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def set_cell(ws, row, col, value, fill=None, font=None, align=None, border=True):
        cell = ws.cell(row=row, column=col, value=value)
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border: cell.border = THIN_BORDER
        return cell

    # --- HOJA 1: INICIO (BLOQUEADA) ---
    ws_inicio = wb.active
    ws_inicio.title = "Inicio"
    
    # Insertar Logo (Fuerza el uso estricto de Logo_enc_fam.png)
    logo_path = "Logo_enc_fam.png"
    if os.path.exists(logo_path):
        try:
            img = OpenpyxlImage(logo_path)
            img.width = 180
            img.height = 90
            ws_inicio.add_image(img, "B2")
        except Exception:
            pass
            
    # Títulos e Instrucciones
    ws_inicio.merge_cells("B7:F7")
    set_cell(ws_inicio, 7, 2, "REPORTE OFICIAL DE GESTIÓN PERCAPITA", None, TITLE_FONT, CENTER, border=False)
    
    ws_inicio.merge_cells("B9:F9")
    set_cell(ws_inicio, 9, 2, "Orientación de la Información", GRAY_FILL, BOLD_DARK, LEFT, border=False)
    
    instrucciones = [
        "Este documento consolida la información de las familias registradas en el sistema ERBI Analytics,",
        "cruzada automáticamente con la base de datos oficial de inscritos (Percapita) del CESFAM Cholchol.",
        "",
        "CONTENIDO DEL REPORTE:",
        "► Pestaña 'Dashboard': Resumen ejecutivo, KPIs y Gráficos estadísticos de la población evaluada.",
        "► Pestaña 'Base de Datos': Listado detallado de integrantes con su estado de percapitación.",
        "",
        "NOTAS DE SEGURIDAD:",
        f"Generado por: {user_info.get('usuario', 'Usuario')} - {user_info.get('cargo', '')}",
        f"Fecha y Hora de emisión: {fecha_generacion} (Hora de Chile)",
        "Esta información es confidencial y para uso exclusivo del equipo clínico y directivo."
    ]
    
    r = 11
    for linea in instrucciones:
        ws_inicio.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        set_cell(ws_inicio, r, 2, linea, None, NORMAL, LEFT, border=False)
        r += 1
        
    # Ajustar anchos y ocultar líneas de cuadrícula
    ws_inicio.column_dimensions['A'].width = 5
    ws_inicio.column_dimensions['B'].width = 15
    ws_inicio.column_dimensions['C'].width = 25
    ws_inicio.column_dimensions['D'].width = 25
    ws_inicio.column_dimensions['E'].width = 15
    ws_inicio.sheet_view.showGridLines = False
    
    # Proteger la hoja
    ws_inicio.protection.sheet = True
    ws_inicio.protection.password = "erbianalytics2026"

    # --- DATOS PARA GRÁFICOS (HOJA OCULTA) ---
    ws_data = wb.create_sheet(title="Datos_Graficos")
    ws_data.sheet_state = 'hidden' # Ocultar para el usuario
    
    total_eval = len(df_cruzado)
    percapitados = len(df_cruzado[df_cruzado['Estado Percapita'] == 'Percapitado'])
    no_percapitados = total_eval - percapitados
    pct_percap = round((percapitados / total_eval * 100), 1) if total_eval > 0 else 0
    
    # Tabla 1: Estado de Percapitación
    ws_data['A1'] = "Estado"
    ws_data['B1'] = "Cantidad"
    ws_data['A2'] = "Percapitado"
    ws_data['B2'] = percapitados
    ws_data['A3'] = "No Percapitado"
    ws_data['B3'] = no_percapitados
    
    # Tabla 2: Distribución por Centro
    ws_data['D1'] = "Centro"
    ws_data['E1'] = "Cantidad"
    df_centros = df_cruzado.groupby('Centro Percapita').size().reset_index(name='count')
    r_data = 2
    for _, row in df_centros.iterrows():
        ws_data.cell(row=r_data, column=4, value=str(row['Centro Percapita']))
        ws_data.cell(row=r_data, column=5, value=row['count'])
        r_data += 1
        
    # Tabla 3: Distribución por Sector
    ws_data['G1'] = "Sector"
    ws_data['H1'] = "Cantidad"
    df_sectores = df_cruzado.groupby('Sector').size().reset_index(name='count')
    r_sec = 2
    for _, row in df_sectores.iterrows():
        ws_data.cell(row=r_sec, column=7, value=str(row['Sector']))
        ws_data.cell(row=r_sec, column=8, value=row['count'])
        r_sec += 1
    
    # --- HOJA 2: DASHBOARD ---
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.sheet_view.showGridLines = False
    
    ws_dash.merge_cells("B2:K3")
    set_cell(ws_dash, 2, 2, "DASHBOARD PERCAPITA", DARK_BLUE, TITLE_FONT, CENTER, border=False)
    ws_dash.cell(row=2, column=2).font = Font(bold=True, color="FFFFFF", size=16)
    
    ws_dash.merge_cells("B5:K5")
    set_cell(ws_dash, 5, 2, f"Periodo de Análisis: {periodo_str}", CELESTE, BOLD_DARK, CENTER, border=False)
    
    # Tarjetas de KPIs
    ws_dash.merge_cells("C7:D7")
    set_cell(ws_dash, 7, 3, "Total Evaluados", GRAY_FILL, BOLD_BLACK, CENTER)
    ws_dash.merge_cells("C8:D9")
    set_cell(ws_dash, 8, 3, total_eval, None, TITLE_FONT, CENTER)
    
    ws_dash.merge_cells("F7:G7")
    set_cell(ws_dash, 7, 6, "Validados (Percapitados)", GREEN_FILL, BOLD_BLACK, CENTER)
    ws_dash.merge_cells("F8:G9")
    set_cell(ws_dash, 8, 6, f"{percapitados} ({pct_percap}%)", None, GREEN_FONT, CENTER)
    ws_dash.cell(row=8, column=6).font = Font(color="166534", bold=True, size=14)
    
    ws_dash.merge_cells("I7:J7")
    set_cell(ws_dash, 7, 9, "No Percapitados", RED_FILL, BOLD_BLACK, CENTER)
    ws_dash.merge_cells("I8:J9")
    set_cell(ws_dash, 8, 9, f"{no_percapitados} ({round(100-pct_percap, 1)}%)", None, RED_FONT, CENTER)
    ws_dash.cell(row=8, column=9).font = Font(color="991B1B", bold=True, size=14)

    # Gráfico 1: Torta de Estado de Percapitación
    pie1 = PieChart()
    labels1 = Reference(ws_data, min_col=1, min_row=2, max_row=3)
    data1 = Reference(ws_data, min_col=2, min_row=1, max_row=3)
    pie1.add_data(data1, titles_from_data=True)
    pie1.set_categories(labels1)
    pie1.title = "Distribución Global Percapita"
    pie1.dataLabels = DataLabelList() 
    pie1.dataLabels.showPercent = True
    pie1.width = 12
    pie1.height = 8
    ws_dash.add_chart(pie1, "B12")
    
    # Gráfico 2: Barras por Centro
    bar2 = BarChart()
    data2 = Reference(ws_data, min_col=5, min_row=1, max_row=r_data-1)
    cats2 = Reference(ws_data, min_col=4, min_row=2, max_row=r_data-1)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.title = "Integrantes Validados por Establecimiento"
    bar2.x_axis.title = "Establecimiento"
    bar2.y_axis.title = "Cantidad"
    bar2.legend = None
    bar2.width = 16
    bar2.height = 8
    ws_dash.add_chart(bar2, "F12")

    # Gráfico 3: Torta de Sector
    pie3 = PieChart()
    labels3 = Reference(ws_data, min_col=7, min_row=2, max_row=r_sec-1)
    data3 = Reference(ws_data, min_col=8, min_row=1, max_row=r_sec-1)
    pie3.add_data(data3, titles_from_data=True)
    pie3.set_categories(labels3)
    pie3.title = "Distribución por Sector"
    pie3.dataLabels = DataLabelList()
    pie3.dataLabels.showPercent = True
    pie3.width = 12
    pie3.height = 8
    ws_dash.add_chart(pie3, "B28")

    # Ajustar anchos
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws_dash.column_dimensions[col].width = 12

    # --- HOJA 3: BASE DE DATOS ---
    ws_db = wb.create_sheet(title="Base de Datos")
    
    # Cabecera Institucional BD
    r = 1
    ws_db.row_dimensions[r].height = 30
    ws_db.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    set_cell(ws_db, r, 1, f"BASE DE DATOS CRUZADA - {periodo_str}", DARK_BLUE, BOLD_WHITE, CENTER, border=False)
    
    r += 2
    # Encabezados de tabla
    headers = list(df_cruzado.columns)
    for i, h in enumerate(headers, 1):
        set_cell(ws_db, r, i, h, DARK_BLUE, BOLD_WHITE, CENTER)
        
    r += 1
    for _, row_data in df_cruzado.iterrows():
        for i, h in enumerate(headers, 1):
            val = str(row_data[h])
            # Color semántico estado
            if h == 'Estado Percapita':
                if val == 'Percapitado':
                    set_cell(ws_db, r, i, val, GREEN_FILL, GREEN_FONT, CENTER)
                else:
                    set_cell(ws_db, r, i, val, RED_FILL, RED_FONT, CENTER)
            else:
                set_cell(ws_db, r, i, val, None, NORMAL, LEFT if i in [2,4] else CENTER)
        r += 1

    # Ajustar anchos BD
    widths = [15, 25, 12, 35, 15, 15, 20]
    for i, w in enumerate(widths, 1):
        ws_db.column_dimensions[get_column_letter(i)].width = w
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
